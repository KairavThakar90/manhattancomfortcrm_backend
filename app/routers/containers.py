"""
Container API endpoints
"""
from typing import Optional, List, Union, Dict, Any, Set, Tuple
from datetime import datetime, timezone, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, BackgroundTasks, Form, Request
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.auth import get_current_user
from app import models, schemas
from app.schemas import (
    ContainerOut, ContainerCreate, POItemsForContainerResponse,
    ContainerListResponse, ContainerDetailOut, ContainerDetailItemOut,
    ContainerUpdate, ContainerWarehouseUpdate, ContainerAddItems, ContainerActivityCreate, ContainerAttachmentOut,
    UserActivityLogOut, PaginatedResponse, ContainerTrackingOut, ShippingContainerCommentOut
)
from app.services.sellercloud_client import SellerCloudClient
from app.services.activity_service import log_activity
from app.services.gcs_service import upload_file_to_gcs

import uuid


def format_sc_date(dt):
    if not dt: return None
    return dt.strftime("%Y-%m-%dT12:00:00Z")

def resolve_container_filter(container_id: str):
    """
    Helper to allow fetching a container by either its local UUID or its SellerCloud integer ID.
    Returns the SQLAlchemy filter condition.
    """
    try:
        container_uuid = uuid.UUID(container_id)
        return models.ShippingContainer.id == container_uuid
    except ValueError:
        if container_id.isdigit():
            return models.ShippingContainer.sellercloud_container_id == int(container_id)
        raise HTTPException(status_code=400, detail="Invalid container ID format (must be UUID or SellerCloud integer ID)")


def resolve_warehouse_helper(
    db: Session,
    warehouse_id: Optional[Union[uuid.UUID, int, str]] = None,
    sellercloud_warehouse_id: Optional[int] = None,
    warehouse_name: Optional[str] = None
) -> Optional[models.Warehouse]:
    """
    Resolves a warehouse by SellerCloud integer ID, local database UUID, or warehouse name.
    """
    if sellercloud_warehouse_id is not None:
        wh = db.query(models.Warehouse).filter(models.Warehouse.sellercloud_warehouse_id == sellercloud_warehouse_id).first()
        if wh:
            return wh

    if warehouse_name:
        wh = db.query(models.Warehouse).filter(models.Warehouse.name.ilike(warehouse_name.strip())).first()
        if wh:
            return wh

    if warehouse_id is not None:
        if isinstance(warehouse_id, int) or (isinstance(warehouse_id, str) and str(warehouse_id).isdigit()):
            wh = db.query(models.Warehouse).filter(models.Warehouse.sellercloud_warehouse_id == int(warehouse_id)).first()
            if wh:
                return wh
        try:
            val_uuid = uuid.UUID(str(warehouse_id))
            wh = db.query(models.Warehouse).filter(models.Warehouse.id == val_uuid).first()
            if wh:
                return wh
        except ValueError:
            wh = db.query(models.Warehouse).filter(models.Warehouse.name.ilike(str(warehouse_id).strip())).first()
            if wh:
                return wh

    return None

router = APIRouter(
    prefix="/containers",
    tags=["Containers"],
    dependencies=[Depends(get_current_user)],
)


def compute_container_status(
    total_qty_in_container: int,
    total_qty_received: int,
    date_emptied: Optional[datetime],
    date_dropped_off: Optional[datetime],
    received_date: Optional[datetime] = None
) -> tuple[str, str]:
    """
    Computes container status enum code and human-readable label.
    Order of evaluation:
    1. Fully Received: total_qty_received >= total_qty_in_container (and > 0), or received_date set when in_container == 0
    2. Partially Received: 0 < total_qty_received < total_qty_in_container
    3. Unloaded / Emptied: date_emptied is set
    4. Picked Up: date_dropped_off is set
    5. In Transit: default (no dates set)
    Returns: (status_code, status_label)
    """
    if (total_qty_in_container > 0 and total_qty_received >= total_qty_in_container) or (total_qty_in_container == 0 and received_date is not None):
        return ("FULLY_RECEIVED", "Fully Received")
    if 0 < total_qty_received < total_qty_in_container:
        return ("PARTIALLY_RECEIVED", "Partially Received")
    if date_emptied is not None:
        return ("UNLOADED_EMPTIED", "Unloaded/Emptied")
    if date_dropped_off is not None:
        return ("PICKED_UP", "Picked Up")
    return ("IN_TRANSIT", "In Transit")


# ---------------------------------------------------------------------------
# GET /containers/  — paginated list with filters
# ---------------------------------------------------------------------------
@router.get("")
@router.get("/")
def list_containers(
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(25, ge=1, le=200, description="Items per page"),
    received: Optional[bool] = Query(None, description="True = received only, False = not received, omit = all"),
    container_status: Optional[str] = Query(None, description="Filter by container status: FULLY_RECEIVED, PARTIALLY_RECEIVED, UNLOADED_EMPTIED, PICKED_UP, IN_TRANSIT (comma-separated for multiple)"),
    status: Optional[str] = Query(None, description="Alias for container_status: FULLY_RECEIVED, PARTIALLY_RECEIVED, UNLOADED_EMPTIED, PICKED_UP, IN_TRANSIT"),
    search: Optional[str] = Query(None, description="Search by container name (partial match)"),
    po_id: Optional[str] = Query(None, description="Filter by Purchase Order UUID"),
    sellercloud_po_id: Optional[int] = Query(None, description="Filter by SellerCloud PO integer ID"),
    vendor_id: Optional[str] = Query(None, description="Filter by vendor UUID"),
    sellercloud_warehouse_id: Optional[int] = Query(None, description="Filter by SellerCloud Warehouse ID"),
    date_from: Optional[datetime] = Query(None, description="Filter containers received on or after this date"),
    date_to: Optional[datetime] = Query(None, description="Filter containers received on or before this date"),
    receive_date_from: Optional[datetime] = Query(None, description="Filter containers received on or after this date (alias for date_from)"),
    receive_date_to: Optional[datetime] = Query(None, description="Filter containers received on or before this date (alias for date_to)"),
    eta_from: Optional[datetime] = Query(None, description="Filter containers with ETA on or after this date"),
    eta_to: Optional[datetime] = Query(None, description="Filter containers with ETA on or before this date"),
    sort_by: Optional[str] = Query(None, description="Sort by: eta_delivery, receive_date, status, date_emptied"),
    sort_order: Optional[str] = Query("desc", description="Sort order: asc or desc"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Paginated list of all shipping containers with filters and summary counts.

    **Filters:**
    - `status` / `container_status` — FULLY_RECEIVED, PARTIALLY_RECEIVED, UNLOADED_EMPTIED, PICKED_UP, IN_TRANSIT
    - `received=true` — only containers with a received date
    - `received=false` — only containers NOT yet received (in transit / pending)
    - `search` — partial container name search (case-insensitive)
    - `po_id` — containers linked to a specific PO (local UUID)
    - `sellercloud_po_id` — containers linked to a specific PO (SC integer ID)
    - `vendor_id` — containers linked to POs from a specific vendor
    - `sellercloud_warehouse_id` — containers stored in a specific warehouse (SC integer ID)
    - `sort_by` — `eta_delivery`, `receive_date`, `date_emptied`, `status`
    - `sort_order` — `asc` or `desc`

    Each result includes:
    - `is_received` — boolean derived from received_date
    - `total_items` — number of distinct SKUs in the container
    - `total_qty_in_container` — total units across all items
    - `total_qty_received` — total units received from SC
    - `unique_pos` — number of distinct POs in the container
    """
    query = db.query(models.ShippingContainer).options(
        joinedload(models.ShippingContainer.warehouse),
        joinedload(models.ShippingContainer.attachments),
        joinedload(models.ShippingContainer.comments),
        joinedload(models.ShippingContainer.logistics_company),
        joinedload(models.ShippingContainer.item_links).joinedload(models.PurchaseOrderItemContainer.item).joinedload(models.PurchaseOrderItem.purchase_order)
    )

    # Received / not-received filter
    if received is True:
        query = query.filter(models.ShippingContainer.received_date.isnot(None))
    elif received is False:
        query = query.filter(models.ShippingContainer.received_date.is_(None))

    # Filter by Container Status (accepts either container_status or status query param)
    effective_status = container_status or status
    if effective_status:
        requested_statuses = [
            s.strip().upper().replace(" ", "_").replace("-", "_").replace("/", "_")
            for s in effective_status.split(",")
            if s.strip()
        ]
        
        from sqlalchemy import func, case, and_, or_
        from app.models import PurchaseOrderItemContainer
        
        agg_sub = (
            db.query(
                PurchaseOrderItemContainer.shipping_container_id.label("c_id"),
                func.coalesce(func.sum(PurchaseOrderItemContainer.qty_in_container), 0).label("tot_in"),
                func.coalesce(func.sum(PurchaseOrderItemContainer.qty_received_container), 0).label("tot_recv"),
            )
            .group_by(PurchaseOrderItemContainer.shipping_container_id)
            .subquery()
        )
        
        tot_in_col = func.coalesce(agg_sub.c.tot_in, 0)
        tot_recv_col = func.coalesce(agg_sub.c.tot_recv, 0)
        
        query = query.outerjoin(agg_sub, models.ShippingContainer.id == agg_sub.c.c_id)
        
        cond_map = {
            "FULLY_RECEIVED": or_(
                and_(tot_in_col > 0, tot_recv_col >= tot_in_col),
                and_(tot_in_col == 0, models.ShippingContainer.received_date.isnot(None))
            ),
            "PARTIALLY_RECEIVED": and_(
                tot_recv_col > 0,
                tot_recv_col < tot_in_col
            ),
            "UNLOADED_EMPTIED": and_(
                models.ShippingContainer.date_emptied.isnot(None),
                tot_recv_col == 0,
                models.ShippingContainer.received_date.is_(None)
            ),
            "PICKED_UP": and_(
                models.ShippingContainer.date_dropped_off.isnot(None),
                models.ShippingContainer.date_emptied.is_(None),
                tot_recv_col == 0,
                models.ShippingContainer.received_date.is_(None)
            ),
            "IN_TRANSIT": and_(
                models.ShippingContainer.date_dropped_off.is_(None),
                models.ShippingContainer.date_emptied.is_(None),
                tot_recv_col == 0,
                models.ShippingContainer.received_date.is_(None)
            ),
        }
        
        filter_exprs = []
        for req in requested_statuses:
            if req in cond_map:
                filter_exprs.append(cond_map[req])
            elif req in ("UNLOADED", "EMPTIED", "UNLOADED_EMPTIED"):
                filter_exprs.append(cond_map["UNLOADED_EMPTIED"])
            elif req in ("PICKED", "PICKEDUP", "PICKED_UP", "DROPPED_OFF", "DROPPEDOFF"):
                filter_exprs.append(cond_map["PICKED_UP"])
            elif req in ("TRANSIT", "INTRANSIT", "IN_TRANSIT"):
                filter_exprs.append(cond_map["IN_TRANSIT"])
            elif req in ("PARTIAL", "PARTIALLY", "PARTIAL_RECEIVED", "PARTIALLY_RECEIVED"):
                filter_exprs.append(cond_map["PARTIALLY_RECEIVED"])
            elif req in ("FULL", "FULLY", "FULL_RECEIVED", "FULLY_RECEIVED", "RECEIVED"):
                filter_exprs.append(cond_map["FULLY_RECEIVED"])

        if filter_exprs:
            query = query.filter(or_(*filter_exprs))

    order_clauses = []

    # Container name or ID search
    if search:
        from sqlalchemy import or_, case, cast, String
        import re
        escaped_search = re.escape(search)
        search_conditions = [
            models.ShippingContainer.container_name.op('~*')(rf"\y{escaped_search}")
        ]
        if search.isdigit():
            search_conditions.append(cast(models.ShippingContainer.sellercloud_container_id, String).op('~*')(rf"\y{escaped_search}"))
            order_clauses.append(case((models.ShippingContainer.sellercloud_container_id == int(search), 0), else_=1))
            
        query = query.filter(or_(*search_conditions))

    # Filter by Warehouse
    if sellercloud_warehouse_id:
        query = query.join(models.Warehouse, models.ShippingContainer.warehouse_id == models.Warehouse.id).filter(
            models.Warehouse.sellercloud_warehouse_id == sellercloud_warehouse_id
        )

    # Restrict Warehouse Users to their assigned warehouse
    if current_user.role == "warehouse" and current_user.warehouse_id:
        query = query.filter(models.ShippingContainer.warehouse_id == current_user.warehouse_id)

    # Restrict Vendors to their assigned vendor containers
    if current_user.role == "vendor" and current_user.vendor_id:
        vendor_container_ids = db.query(models.PurchaseOrderItemContainer.shipping_container_id) \
            .join(models.PurchaseOrderItem, models.PurchaseOrderItemContainer.purchase_order_item_id == models.PurchaseOrderItem.id) \
            .join(models.PurchaseOrder, models.PurchaseOrderItem.purchase_order_id == models.PurchaseOrder.id) \
            .filter(models.PurchaseOrder.vendor_id == current_user.vendor_id)
        query = query.filter(models.ShippingContainer.id.in_(vendor_container_ids))

    # Filter by PO (UUID or SC integer ID)
    if po_id or sellercloud_po_id or vendor_id:
        po_query = db.query(models.PurchaseOrderItemContainer.shipping_container_id) \
            .join(models.PurchaseOrderItem, models.PurchaseOrderItemContainer.purchase_order_item_id == models.PurchaseOrderItem.id) \
            .join(models.PurchaseOrder, models.PurchaseOrderItem.purchase_order_id == models.PurchaseOrder.id)
            
        if po_id:
            po_query = po_query.filter(models.PurchaseOrder.id == po_id)
        if sellercloud_po_id:
            po_query = po_query.filter(models.PurchaseOrder.sellercloud_po_id == sellercloud_po_id)
        if vendor_id:
            po_query = po_query.filter(models.PurchaseOrder.vendor_id == vendor_id)
            
        query = query.filter(models.ShippingContainer.id.in_(po_query))

    # Filter by received date range (supports both date_from/date_to and receive_date_from/receive_date_to)
    effective_date_from = date_from or receive_date_from
    effective_date_to = date_to or receive_date_to

    if effective_date_from:
        query = query.filter(models.ShippingContainer.received_date >= effective_date_from)
    if effective_date_to:
        # If effective_date_to has no time component (midnight), extend it to the end of the day
        if effective_date_to.time() == datetime.min.time():
            from datetime import time
            effective_date_to = datetime.combine(effective_date_to.date(), time(23, 59, 59, 999999))
        query = query.filter(models.ShippingContainer.received_date <= effective_date_to)

    # Filter by ETA date range
    if eta_from:
        query = query.filter(models.ShippingContainer.estimated_arrival_date >= eta_from)
    if eta_to:
        # If eta_to has no time component (midnight), extend it to the end of the day
        if eta_to.time() == datetime.min.time():
            from datetime import time
            eta_to = datetime.combine(eta_to.date(), time(23, 59, 59, 999999))
        query = query.filter(models.ShippingContainer.estimated_arrival_date <= eta_to)

    sort_by_clean = (sort_by or "").lower().replace("-", "_").replace(" ", "_").strip()
    is_asc = sort_order and str(sort_order).lower().strip() in ("asc", "ascending", "1")

    if sort_by_clean in ("eta_delivery", "etadelivery", "eta", "estimated_arrival_date"):
        sort_col = models.ShippingContainer.estimated_arrival_date
    elif sort_by_clean in ("receive_date", "receivedate", "received_date", "receive"):
        sort_col = models.ShippingContainer.received_date
    elif sort_by_clean in ("date_emptied", "dateemptied", "emptied_date", "emptieddate", "emptied", "empty_date"):
        sort_col = models.ShippingContainer.date_emptied
    elif sort_by_clean in ("date_dropped_off", "datedroppedoff", "dropped_off_date", "dropped_off"):
        sort_col = models.ShippingContainer.date_dropped_off
    elif sort_by_clean in ("status", "is_received"):
        sort_col = models.ShippingContainer.received_date.is_(None)
    elif sort_by_clean in ("container_name", "name", "container"):
        sort_col = models.ShippingContainer.container_name
    elif sort_by_clean in ("created_at", "createdat", "created"):
        sort_col = models.ShippingContainer.created_at
    else:
        sort_col = None

    order_clauses = []
    if sort_by_clean in ("date_emptied", "dateemptied", "emptied_date", "emptieddate", "emptied", "empty_date", "is_emptied", "has_date_emptied"):
        if is_asc:
            # Non-emptied first (date_emptied IS NULL), then emptied
            order_clauses.append(models.ShippingContainer.date_emptied.isnot(None).asc())
            order_clauses.append(models.ShippingContainer.date_emptied.asc().nullslast())
            order_clauses.append(models.ShippingContainer.created_at.desc())
        else:
            # Emptied containers first (date_emptied IS NOT NULL), then non-emptied
            order_clauses.append(models.ShippingContainer.date_emptied.isnot(None).desc())
            order_clauses.append(models.ShippingContainer.date_emptied.desc().nullslast())
            order_clauses.append(models.ShippingContainer.created_at.desc())
    elif sort_col is not None:
        if is_asc:
            order_clauses.append(sort_col.asc().nullslast())
            order_clauses.append(models.ShippingContainer.created_at.asc())
        else:
            order_clauses.append(sort_col.desc().nullslast())
            order_clauses.append(models.ShippingContainer.created_at.desc())
    else:
        order_clauses.append(models.ShippingContainer.created_at.desc())

    total = query.count()
    containers = (
        query.order_by(*order_clauses)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    # Build enriched results with summary counts
    results = []
    for ctr in containers:
        links = ctr.item_links  # loaded via relationship
        total_items = len(links)
        total_qty = sum(lnk.qty_in_container or 0 for lnk in links)
        total_received = sum(lnk.qty_received_container or 0 for lnk in links)
        unique_po_ids = set()
        po_numbers_set = set()
        for lnk in links:
            if lnk.item and lnk.item.purchase_order_id:
                unique_po_ids.add(lnk.item.purchase_order_id)
                if getattr(lnk.item, "purchase_order", None) and lnk.item.purchase_order.sellercloud_po_id:
                    po_numbers_set.add(lnk.item.purchase_order.sellercloud_po_id)

        receiving_count = sum(1 for c in ctr.comments if c.category == "receiving_closure")
        vendor_count = sum(1 for c in ctr.comments if c.category == "vendor_credit")
        total_comments = len(ctr.comments)

        status_code, status_label = compute_container_status(
            total_qty_in_container=total_qty,
            total_qty_received=total_received,
            date_emptied=ctr.date_emptied,
            date_dropped_off=ctr.date_dropped_off,
            received_date=ctr.received_date
        )

        results.append(
            ContainerOut(
                id=ctr.id,
                sellercloud_container_id=ctr.sellercloud_container_id,
                container_name=ctr.container_name,
                estimated_arrival_date=ctr.estimated_arrival_date,
                received_date=ctr.received_date,
                is_received=ctr.received_date is not None,
                container_status=status_code,
                container_status_label=status_label,
                warehouse_id=ctr.warehouse_id,
                warehouse=ctr.warehouse,
                created_at=ctr.created_at,
                updated_at=ctr.updated_at,
                date_dropped_off=ctr.date_dropped_off,
                door=ctr.door,
                trucker_email=ctr.trucker_email,
                trucking_company=ctr.trucking_company,
                logistics_company_id=ctr.logistics_company_id,
                logistics_company=ctr.logistics_company,
                date_emptied=ctr.date_emptied,
                unloaded_by=ctr.unloaded_by,
                unload_cost=float(ctr.unload_cost) if ctr.unload_cost is not None else None,
                container_shipping_cost=float(ctr.container_shipping_cost) if ctr.container_shipping_cost is not None else None,
                drayage_cost=float(ctr.drayage_cost) if ctr.drayage_cost is not None else None,
                customs_duty_misc=float(ctr.customs_duty_misc) if ctr.customs_duty_misc is not None else None,
                per_diem=float(ctr.per_diem) if ctr.per_diem is not None else None,
                country_of_origin=ctr.country_of_origin,
                receiving_closure_notes=ctr.receiving_closure_notes,
                factory_credit_needed=ctr.factory_credit_needed,
                total_items=total_items,
                total_qty_in_container=total_qty,
                total_qty_received=total_received,
                unique_pos=len(unique_po_ids),
                po_numbers=list(po_numbers_set),
                attachments=[ContainerAttachmentOut.model_validate(a) for a in ctr.attachments],
                receiving_closure_comment_count=receiving_count,
                vendor_credit_comment_count=vendor_count,
                comments_count=total_comments
            )
        )

    total_pages = (total + page_size - 1) // page_size if page_size > 0 else 0
    return ContainerListResponse(
        total=total,
        page=page,
        page_size=page_size,
        meta={
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "has_next": page * page_size < total,
            "has_prev": page > 1,
        },
        results=results,
    )


# ---------------------------------------------------------------------------
# GET /containers/po-items/{sellercloud_po_id}
# Returns PO line-items in a shape ready for the "create container" form.
# ---------------------------------------------------------------------------
@router.get("/po-items/{sellercloud_po_id}", response_model=POItemsForContainerResponse)
def get_po_items_for_container(
    sellercloud_po_id: int,
    db: Session = Depends(get_db),
):
    """
    Get all items from a PO formatted for the container-creation workflow.

    **For each line item returns:**
    - `po_item_id` — UUID to use in `POST /containers`
    - `sellercloud_item_id`, `sku`, `product_name`
    - `qty_ordered`, `qty_received`, `qty_remaining`
    - `qty_already_in_containers` — total already assigned to any container
    - `qty_available_for_container` — how many can still be added (ordered − already)
    - `existing_containers` — containers this item is already part of

    **Typical usage:**
    1. `GET /api/v1/containers/po-items/12345` — get item list
    2. User picks qty per item
    3. `POST /api/v1/containers` with the filled payload
    """
    po = (
        db.query(models.PurchaseOrder)
        .options(
            joinedload(models.PurchaseOrder.vendor),
            joinedload(models.PurchaseOrder.items)
            .joinedload(models.PurchaseOrderItem.container_links)
            .joinedload(models.PurchaseOrderItemContainer.container),
        )
        .filter(models.PurchaseOrder.sellercloud_po_id == sellercloud_po_id)
        .first()
    )

    if not po:
        raise HTTPException(
            status_code=404, detail=f"Purchase Order {sellercloud_po_id} not found"
        )

    from app.schemas import POItemForContainerOut, ContainerSummary as CS

    items_out = []
    for item in po.items:
        qty_already = sum(link.qty_in_container or 0 for link in item.container_links)
        qty_available = max(0, item.qty_ordered - qty_already)

        existing_containers = [
            CS(
                id=link.container.id,
                sellercloud_container_id=link.container.sellercloud_container_id,
                container_name=link.container.container_name,
                estimated_arrival_date=link.container.estimated_arrival_date,
                received_date=link.container.received_date,
                qty_in_container=link.qty_in_container,
            )
            for link in item.container_links
            if link.container
        ]

        items_out.append(
            POItemForContainerOut(
                po_item_id=item.id,
                sellercloud_item_id=item.sellercloud_item_id,
                sellercloud_po_id=po.sellercloud_po_id,
                sku=item.sku,
                product_name=item.product_name,
                image_url=item.image_url,
                qty_ordered=item.qty_ordered,
                qty_received=item.qty_received,
                qty_remaining=qty_available,
                qty_already_in_containers=qty_already,
                qty_available_for_container=qty_available,
                existing_containers=existing_containers,
            )
        )

    total_ordered = sum(i.qty_ordered for i in items_out)
    total_received = sum(i.qty_received for i in items_out)
    total_in_containers = sum(i.qty_already_in_containers for i in items_out)
    total_available = sum(i.qty_available_for_container for i in items_out)

    return POItemsForContainerResponse(
        po_id=po.id,
        sellercloud_po_id=po.sellercloud_po_id,
        po_title=po.purchase_title,
        vendor_name=po.vendor.name if po.vendor else None,
        items=items_out,
        summary={
            "total_items": len(items_out),
            "total_qty_ordered": total_ordered,
            "total_qty_received": total_received,
            "total_qty_in_containers": total_in_containers,
            "total_qty_available_for_container": total_available,
        },
    )


# ---------------------------------------------------------------------------
# POST /containers  (and legacy /containers/create)
# Creates a container locally + syncs to SellerCloud
# ---------------------------------------------------------------------------
@router.post("", status_code=201)
@router.post("/create", status_code=201, include_in_schema=False)  # backward compat
def create_container(
    container_data: ContainerCreate,
    db: Session = Depends(get_db),
):
    """
    Create a new shipping container, sync it to SellerCloud, and link PO items.

    **Workflow:**
    1. `GET /api/v1/containers/po-items/{sellercloud_po_id}` — fetch items + available qty
    2. Build the request body with the items you want in this container
    3. POST here

    **Request body example:**
    ```json
    {
      "container_name": "CONT-2026-001",
      "estimated_arrival_date": "2026-09-01T00:00:00Z",
      "items": [
        { "po_item_id": "<uuid>", "qty_in_container": 50 },
        { "po_item_id": "<uuid>", "qty_in_container": 30 }
      ]
    }
    ```

    **Item resolution order** (use whichever identifiers you have):
    1. `po_item_id` — local UUID (preferred, from po-items endpoint)
    2. `sellercloud_item_id` — SellerCloud PO line item ID
    3. `sellercloud_po_id` + `sku` — PO ID + product SKU

    **What this does:**
    - Validates items exist and qty doesn't exceed what's available
    - Creates container in SellerCloud via `POST /api/ShippingContainers`
    - Saves container + item links in local DB
    - Updates `qty_in_container` on each PO item
    - Returns created container with SC ID (if sync succeeded)
    """
    resolved_items = []

    for item_data in container_data.items:
        item = None

        if item_data.po_item_id:
            item = (
                db.query(models.PurchaseOrderItem)
                .filter(models.PurchaseOrderItem.id == item_data.po_item_id)
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )
        elif item_data.sellercloud_item_id:
            item = (
                db.query(models.PurchaseOrderItem)
                .filter(
                    models.PurchaseOrderItem.sellercloud_item_id
                    == item_data.sellercloud_item_id
                )
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )
        elif item_data.sellercloud_po_id and item_data.sku:
            item = (
                db.query(models.PurchaseOrderItem)
                .join(models.PurchaseOrder)
                .filter(
                    models.PurchaseOrder.sellercloud_po_id == item_data.sellercloud_po_id,
                    models.PurchaseOrderItem.sku == item_data.sku,
                )
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )

        if not item:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"PO item not found — tried: "
                    f"po_item_id={item_data.po_item_id}, "
                    f"sellercloud_item_id={item_data.sellercloud_item_id}, "
                    f"sellercloud_po_id={item_data.sellercloud_po_id}, "
                    f"sku={item_data.sku}"
                ),
            )

        # Guard: don't allow over-assignment beyond qty_ordered
        qty_rows = (
            db.query(models.PurchaseOrderItemContainer.qty_in_container)
            .filter(
                models.PurchaseOrderItemContainer.purchase_order_item_id == item.id
            )
            .all()
        )
        qty_already_total = sum(r[0] or 0 for r in qty_rows)
        qty_available = max(0, item.qty_ordered - qty_already_total)

        if item_data.qty_in_container > qty_available:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"SKU '{item.sku}': requested {item_data.qty_in_container} "
                    f"but only {qty_available} available "
                    f"(ordered={item.qty_ordered}, already_in_containers={qty_already_total})"
                ),
            )

        resolved_items.append((item_data, item))

    # Build SellerCloud create payload — SC only accepts name/dates at creation time
    # Items are added in a SECOND call after the container is created
    sc_container_payload = {
        "ContainerName": container_data.container_name,
        "EstimatedArrivalDate": format_sc_date(container_data.estimated_arrival_date),
        "ShippingStatus": 2 if container_data.received_date else 1,
    }
    if container_data.received_date:
        sc_container_payload["ReceivedDate"] = format_sc_date(container_data.received_date)
        
    if container_data.warehouse_id:
        import uuid
        if isinstance(container_data.warehouse_id, uuid.UUID) or (isinstance(container_data.warehouse_id, str) and '-' in str(container_data.warehouse_id)):
            warehouse = db.query(models.Warehouse).filter(models.Warehouse.id == str(container_data.warehouse_id)).first()
        else:
            warehouse = db.query(models.Warehouse).filter(models.Warehouse.sellercloud_warehouse_id == int(container_data.warehouse_id)).first()
            
        if warehouse:
            if warehouse.sellercloud_warehouse_id:
                sc_container_payload["ReceiveWarehouseID"] = warehouse.sellercloud_warehouse_id
            # Reassign container_data.warehouse_id to the UUID for local saving
            container_data.warehouse_id = warehouse.id

    sc_response: dict = {}
    sellercloud_container_id = None
    sc_sync_error = None
    sc_items_error = None

    try:
        sc_client = SellerCloudClient()

        # STEP 1: Create the container (SC API only accepts name/dates — NOT items)
        try:
            sc_response = sc_client.create_shipping_container(sc_container_payload)
            if isinstance(sc_response, dict):
                sellercloud_container_id = (
                    sc_response.get("ID")
                    or sc_response.get("Id")
                    or sc_response.get("ContainerID")
                    or sc_response.get("ContainerId")
                    or sc_response.get("id")
                )
        except Exception as create_err:
            print(f"[create_container] SC Create failed (maybe duplicate name): {create_err}")
            sc_response = {"error": str(create_err)}

        # Fallback: SC sometimes returns empty body — search by name to recover ID
        if not sellercloud_container_id:
            try:
                search_resp = sc_client.search_containers_by_name(container_data.container_name)
                matches = search_resp.get("Items") or []
                for m in matches:
                    if m.get("ContainerName") == container_data.container_name:
                        sellercloud_container_id = m.get("ID") or m.get("Id")
                        if sellercloud_container_id:
                            sc_response["_recovered_id"] = sellercloud_container_id
                            break
            except Exception as lookup_err:
                print(f"[create_container] ID recovery search failed: {lookup_err}")

        # STEP 2: Add items to the container (separate SC endpoint)
        if sellercloud_container_id:
            try:
                sc_items = [
                    {
                        "PurchaseOrderID": (
                            item.purchase_order.sellercloud_po_id
                            if item.purchase_order and item.purchase_order.sellercloud_po_id
                            else item_data.sellercloud_po_id
                        ),
                        "PurchaseOrderItemID": item.sellercloud_item_id or item_data.sellercloud_item_id,
                        "Qty": item_data.qty_in_container,
                    }
                    for item_data, item in resolved_items
                ]
                sc_client.add_items_to_container(sellercloud_container_id, sc_items)
                print(f"[create_container] Added {len(sc_items)} items to SC container {sellercloud_container_id}")
            except Exception as items_err:
                sc_items_error = str(items_err)
                print(f"[create_container] SC add-items failed (container created, items not linked in SC): {sc_items_error}")
                sc_response["_items_error"] = sc_items_error

    except Exception as exc:
        sc_sync_error = str(exc)
        print(f"[create_container] SellerCloud sync failed — saving locally anyway: {exc}")

    # Persist locally regardless of SC outcome
    new_container = models.ShippingContainer(
        container_name=container_data.container_name,
        estimated_arrival_date=container_data.estimated_arrival_date,
        received_date=container_data.received_date,
        sellercloud_container_id=sellercloud_container_id,
        warehouse_id=container_data.warehouse_id,
        date_dropped_off=container_data.date_dropped_off,
        door=container_data.door,
        date_emptied=container_data.date_emptied,
        unloaded_by=container_data.unloaded_by,
        unload_cost=container_data.unload_cost,
        container_shipping_cost=container_data.container_shipping_cost,
        drayage_cost=container_data.drayage_cost,
        customs_duty_misc=container_data.customs_duty_misc,
        per_diem=container_data.per_diem,
        country_of_origin=container_data.country_of_origin,
        receiving_closure_notes=container_data.receiving_closure_notes,
        factory_credit_needed=container_data.factory_credit_needed,
        raw_json=sc_response if isinstance(sc_response, dict) else {"error": sc_sync_error},
    )
    db.add(new_container)
    db.flush()

    linked_items_summary = []
    for item_data, item in resolved_items:
        db.add(
            models.PurchaseOrderItemContainer(
                purchase_order_item_id=item.id,
                shipping_container_id=new_container.id,
                qty_in_container=item_data.qty_in_container,
                raw_json={"sc_payload_item": item_data.model_dump(mode="json")},
            )
        )
        item.qty_in_container = (item.qty_in_container or 0) + item_data.qty_in_container

        linked_items_summary.append({
            "po_item_id": str(item.id),
            "sellercloud_item_id": item.sellercloud_item_id,
            "sellercloud_po_id": (
                item.purchase_order.sellercloud_po_id if item.purchase_order else None
            ),
            "sku": item.sku,
            "product_name": item.product_name,
            "qty_in_container": item_data.qty_in_container,
            "total_item_qty_in_containers": item.qty_in_container,
        })

    db.commit()
    db.refresh(new_container)

    sc_items_synced = sellercloud_container_id is not None and not sc_items_error

    if not sellercloud_container_id:
        message = "Container saved locally — SellerCloud sync failed (see sc_sync_error)"
    elif sc_items_error:
        message = "Container created in SellerCloud, but items failed to sync (see sc_items_error) — vendor/items will be missing in SellerCloud until this is retried"
    else:
        message = "Container created and synced to SellerCloud"

    response = {
        "success": True,
        "sellercloud_synced": sellercloud_container_id is not None,
        "sellercloud_items_synced": sc_items_synced,
        "message": message,
        "container": {
            "id": str(new_container.id),
            "sellercloud_container_id": new_container.sellercloud_container_id,
            "container_name": new_container.container_name,
            "estimated_arrival_date": (
                new_container.estimated_arrival_date.isoformat()
                if new_container.estimated_arrival_date
                else None
            ),
            "received_date": (
                new_container.received_date.isoformat()
                if new_container.received_date
                else None
            ),
            "items_linked": len(linked_items_summary),
            "items": linked_items_summary,
        },
    }
    if sc_sync_error:
        response["sc_sync_error"] = sc_sync_error
    if sc_items_error:
        response["sc_items_error"] = sc_items_error

    return response


def parse_mentions(text: str, db: Session) -> list[models.User]:
    if not text:
        return []
    
    import re
    # 1. Match email addresses first (e.g. @sanjay.storetransform@gmail.com)
    email_matches = re.findall(r'@([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', text)
    
    # 2. Match names (e.g. @Sanjay_Thakar or @Sanjay)
    name_matches = re.findall(r'@([A-Za-z0-9_\-\.]+)', text)
    
    matched_user_ids = set()
    
    # Try to match emails
    if email_matches:
        users = db.query(models.User).filter(
            models.User.email.in_(email_matches),
            models.User.is_active == True,
            models.User.email != "googlecloudcron@manhattancomfort.com"
        ).all()
        for u in users:
            matched_user_ids.add(u.id)
            
    # Try to match names/usernames
    if name_matches:
        users = db.query(models.User).filter(
            models.User.is_active == True,
            models.User.email != "googlecloudcron@manhattancomfort.com"
        ).all()
        for match in name_matches:
            match_clean = match.lower().replace("_", " ")
            for u in users:
                full_name = (u.full_name or "").lower()
                first_name = (u.first_name or "").lower()
                if full_name and match_clean in full_name:
                    matched_user_ids.add(u.id)
                elif first_name and match_clean in first_name:
                    matched_user_ids.add(u.id)
                    
    if not matched_user_ids:
        return []
        
    return db.query(models.User).filter(models.User.id.in_(list(matched_user_ids))).all()


# ---------------------------------------------------------------------------
# PUT /containers/{container_id}
# ---------------------------------------------------------------------------
@router.put("/{container_id}")
@router.patch("/{container_id}")
async def update_container(
    container_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
    container_data: Optional[str] = Form(None),
    files: List[UploadFile] = File(default=[])
):
    if current_user.role == "vendor":
        raise HTTPException(status_code=403, detail="Vendors cannot update container details")
    import json
    content_type = request.headers.get("content-type", "")
    if "application/json" in content_type:
        try:
            body_dict = await request.json()
            update_data = ContainerUpdate(**body_dict)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Invalid JSON payload: {e}")
    elif container_data:
        try:
            update_data = ContainerUpdate(**json.loads(container_data))
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Invalid JSON in container_data: {e}")
    else:
        try:
            form = await request.form()
            if "container_data" in form:
                update_data = ContainerUpdate(**json.loads(form["container_data"]))
            elif form:
                form_dict = {k: v for k, v in form.items() if k != "files"}
                update_data = ContainerUpdate(**form_dict)
            else:
                body_bytes = await request.body()
                if body_bytes:
                    update_data = ContainerUpdate(**json.loads(body_bytes.decode("utf-8")))
                else:
                    update_data = ContainerUpdate()
        except Exception:
            update_data = ContainerUpdate()
    """
    Update a container's name, estimated arrival date, or received date.
    Syncs the update to SellerCloud and saves it locally.
    """
    container = db.query(models.ShippingContainer).filter(resolve_container_filter(container_id)).first()
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    # If it's synced to SellerCloud, push the update
    if container.sellercloud_container_id:
        try:
            sc_client = SellerCloudClient()
            final_received_date = update_data.received_date if update_data.received_date is not None else container.received_date
            sc_payload = {
                "ContainerName": update_data.container_name or container.container_name,
                "EstimatedArrivalDate": format_sc_date(update_data.estimated_arrival_date if update_data.estimated_arrival_date else container.estimated_arrival_date),
                "ReceivedDate": format_sc_date(final_received_date),
                "ShippingStatus": 2 if final_received_date else 1,
            }
            sc_client.update_shipping_container(container.sellercloud_container_id, sc_payload)
        except Exception as exc:
            print(f"Failed to update container {container.sellercloud_container_id} in SellerCloud: {exc}")
            # We can optionally fail here or continue saving locally
            # raise HTTPException(status_code=500, detail=f"SellerCloud sync failed: {exc}")

    # Update local DB and track exact changes
    changes = []
    
    if update_data.container_name is not None and update_data.container_name != container.container_name:
        changes.append({"field": "container_name", "old": container.container_name, "new": update_data.container_name})
        container.container_name = update_data.container_name
        
    if update_data.estimated_arrival_date is not None:
        old_val = container.estimated_arrival_date.isoformat() if container.estimated_arrival_date else None
        new_val = update_data.estimated_arrival_date.isoformat() if update_data.estimated_arrival_date else None
        if old_val != new_val:
            changes.append({"field": "estimated_arrival_date", "old": old_val, "new": new_val})
            container.estimated_arrival_date = update_data.estimated_arrival_date
            
    if update_data.received_date is not None:
        old_val = container.received_date.isoformat() if container.received_date else None
        new_val = update_data.received_date.isoformat() if update_data.received_date else None
        if old_val != new_val:
            changes.append({"field": "received_date", "old": old_val, "new": new_val})
            container.received_date = update_data.received_date

    # Update lifecycle fields dynamically if they are passed in the request
    update_dict = update_data.model_dump(exclude_unset=True)
    lifecycle_fields = [
        "date_dropped_off", "door", "date_emptied", "unloaded_by", 
        "unload_cost", "container_shipping_cost", "drayage_cost", "customs_duty_misc", 
        "per_diem", "country_of_origin", "trucker_email", "trucking_company",
        "logistics_company_id"
    ]
    for field in lifecycle_fields:
        if field in update_dict:
            new_val = update_dict[field]
            old_val = getattr(container, field)
            
            import uuid
            if isinstance(old_val, uuid.UUID):
                old_val_str = str(old_val)
            elif isinstance(old_val, datetime):
                old_val_str = old_val.isoformat()
            else:
                old_val_str = old_val
                
            if isinstance(new_val, uuid.UUID):
                new_val_str = str(new_val)
            elif isinstance(new_val, datetime):
                new_val_str = new_val.isoformat()
            else:
                new_val_str = new_val
                
            def _normalize(v):
                return None if v == "" else v

            if _normalize(old_val_str) != _normalize(new_val_str):
                changes.append({"field": field, "old": old_val_str, "new": new_val_str})
                
            setattr(container, field, new_val)

    # Process files
    uploaded_attachments = []
    if files:
        allowed_types = [
            "application/pdf",
            "application/msword",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "text/csv"
        ]
        for f in files:
            if f.size and f.size > 5 * 1024 * 1024:
                raise HTTPException(status_code=400, detail=f"File {f.filename} exceeds 5MB limit.")
            if f.filename and f.content_type:
                if not f.content_type.startswith('image/') and f.content_type not in allowed_types:
                    raise HTTPException(status_code=400, detail=f"File type not allowed for {f.filename}.")
                    
        for f in files:
            if not f.filename:
                continue
            file_bytes = await f.read()
            if len(file_bytes) > 5 * 1024 * 1024:
                raise HTTPException(status_code=400, detail=f"File {f.filename} exceeds 5MB limit.")
            file_url = await upload_file_to_gcs(file_bytes, f.filename, f.content_type)
            att_model = models.ShippingContainerAttachment(
                shipping_container_id=container.id,
                file_name=f.filename,
                file_url=file_url,
                content_type=f.content_type,
                size=len(file_bytes)
            )
            db.add(att_model)
            uploaded_attachments.append(att_model)

    # Explicitly load logistics company relationship if ID is present
    if container.logistics_company_id and not container.logistics_company:
        container.logistics_company = db.query(models.LogisticsCompany).filter(models.LogisticsCompany.id == container.logistics_company_id).first()

    # Auto-populate trucker_email if it's empty but we have a logistics company primary email
    if not container.trucker_email and container.logistics_company and container.logistics_company.primary_email:
        container.trucker_email = container.logistics_company.primary_email

    container.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(container)

    # Email notification logic
    if container.date_emptied:
        email_to = None
        if container.logistics_company and container.logistics_company.primary_email:
            email_to = container.logistics_company.primary_email
        else:
            email_to = container.trucker_email

        if email_to and email_to != container.last_notified_trucker_email:
            from app.services.email_service import send_container_emptied_notification
            formatted_date = container.date_emptied.strftime('%Y-%m-%d')
            formatted_dropped_date = container.date_dropped_off.strftime('%Y-%m-%d') if container.date_dropped_off else None
            
            # Base CC list: Internal CRM users who enabled notifications
            users_to_notify = db.query(models.User).filter(models.User.notify_trucker_email == True).all()
            cc_emails = [u.email for u in users_to_notify if u.email]
            
            # Append Logistics Company CC emails if present
            if container.logistics_company and container.logistics_company.cc_email:
                # cc_email can be comma-separated, split and clean
                extra_ccs = [email.strip() for email in container.logistics_company.cc_email.split(',') if email.strip()]
                for email in extra_ccs:
                    if email not in cc_emails:
                        cc_emails.append(email)
            
            background_tasks.add_task(
                send_container_emptied_notification,
                email_to=email_to,
                container_name=container.container_name,
                date_emptied=formatted_date,
                door_name=container.door,
                date_dropped_off=formatted_dropped_date,
                cc_emails=cc_emails
            )
            container.last_notified_trucker_email = email_to
            db.commit()

    # Legacy tag notifications on save container removed (transitioned to threaded comments)
    details_payload = update_data.model_dump(mode='json', exclude_unset=True)
    if changes:
        details_payload["changes"] = changes
        
    log_activity(db, action="UPDATE_CONTAINER", user_id=current_user.id, entity_type="CONTAINER", entity_id=str(container.id), details=details_payload)

    return {
        "success": True,
        "message": "Container updated successfully",
        "container": {
            "id": str(container.id),
            "container_name": container.container_name,
            "estimated_arrival_date": container.estimated_arrival_date,
            "received_date": container.received_date,
            "sellercloud_container_id": container.sellercloud_container_id,
            "date_dropped_off": container.date_dropped_off,
            "door": container.door,
            "trucker_email": container.trucker_email,
            "trucking_company": container.trucking_company,
            "logistics_company_id": str(container.logistics_company_id) if container.logistics_company_id else None,
            "date_emptied": container.date_emptied,
            "unloaded_by": container.unloaded_by,
            "unload_cost": float(container.unload_cost) if container.unload_cost is not None else None,
            "container_shipping_cost": float(container.container_shipping_cost) if container.container_shipping_cost is not None else None,
            "drayage_cost": float(container.drayage_cost) if container.drayage_cost is not None else None,
            "customs_duty_misc": float(container.customs_duty_misc) if container.customs_duty_misc is not None else None,
            "per_diem": float(container.per_diem) if container.per_diem is not None else None,
            "country_of_origin": container.country_of_origin,
            "receiving_closure_notes": container.receiving_closure_notes,
            "factory_credit_needed": container.factory_credit_needed,
            "attachments": [
                {
                    "id": str(att.id),
                    "file_name": att.file_name,
                    "file_url": att.file_url,
                    "content_type": att.content_type,
                    "size": att.size
                } for att in container.attachments
            ]
        }
    }


# ---------------------------------------------------------------------------
# PUT /containers/{container_id}/warehouse
# ---------------------------------------------------------------------------
@router.put("/{container_id}/warehouse")
@router.patch("/{container_id}/warehouse")
def update_container_warehouse(
    container_id: str,
    payload: ContainerWarehouseUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Update receiving warehouse for a container in both SellerCloud and local database.
    Required parameter: warehouse_id (accepts warehouse UUID, SellerCloud integer ID, or warehouse name).
    """
    if current_user.role == "vendor":
        raise HTTPException(status_code=403, detail="Vendors cannot update container warehouse")
        
    container = db.query(models.ShippingContainer).filter(resolve_container_filter(container_id)).first()
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    warehouse = resolve_warehouse_helper(
        db,
        warehouse_id=payload.warehouse_id,
        sellercloud_warehouse_id=payload.sellercloud_warehouse_id,
        warehouse_name=payload.warehouse_name
    )
    if not warehouse:
        raise HTTPException(
            status_code=400,
            detail="Warehouse not found. Please provide a valid warehouse_id, sellercloud_warehouse_id, or warehouse_name."
        )

    old_wh_name = container.warehouse.name if container.warehouse else None
    sc_updated = False

    # Sync to SellerCloud if container has a SellerCloud ID
    if container.sellercloud_container_id and warehouse.sellercloud_warehouse_id:
        try:
            sc_client = SellerCloudClient()
            sc_resp = sc_client.update_shipping_container(
                container.sellercloud_container_id,
                {
                    "ContainerName": container.container_name,
                    "ReceivingWarehouseID": warehouse.sellercloud_warehouse_id,
                    "EstimatedArrivalDate": format_sc_date(container.estimated_arrival_date),
                    "ShippingStatus": 2 if container.received_date else 1
                }
            )
            sc_updated = sc_resp.get("success", False)
        except Exception as exc:
            print(f"Failed to update warehouse for container {container.sellercloud_container_id} in SellerCloud: {exc}")

    # Update local database
    container.warehouse_id = warehouse.id
    container.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(container)

    log_activity(
        db,
        action="UPDATE_CONTAINER_WAREHOUSE",
        user_id=current_user.id,
        entity_type="CONTAINER",
        entity_id=str(container.id),
        details={
            "old_warehouse": old_wh_name,
            "new_warehouse": warehouse.name,
            "sellercloud_warehouse_id": warehouse.sellercloud_warehouse_id,
            "sellercloud_synced": sc_updated
        }
    )

    return {
        "success": True,
        "message": f"Successfully updated container warehouse to '{warehouse.name}'",
        "container_id": str(container.id),
        "sellercloud_container_id": container.sellercloud_container_id,
        "sellercloud_synced": sc_updated,
        "warehouse": {
            "id": str(warehouse.id),
            "name": warehouse.name,
            "sellercloud_warehouse_id": warehouse.sellercloud_warehouse_id
        }
    }


@router.post("/preview-sc-payload")
def preview_sc_payload(
    container_data: ContainerCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    **DEBUG ONLY** - Resolves items and returns the exact JSON payload 
    that would be sent to SellerCloud's `POST /api/ShippingContainers/{id}/Items` endpoint.
    It does not call SellerCloud and does not modify the database.
    """
    resolved_items = []

    for item_data in container_data.items:
        item = None

        if item_data.po_item_id:
            item = (
                db.query(models.PurchaseOrderItem)
                .filter(models.PurchaseOrderItem.id == item_data.po_item_id)
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )
        elif item_data.sellercloud_item_id:
            item = (
                db.query(models.PurchaseOrderItem)
                .filter(
                    models.PurchaseOrderItem.sellercloud_item_id
                    == item_data.sellercloud_item_id
                )
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )
        elif item_data.sellercloud_po_id and item_data.sku:
            item = (
                db.query(models.PurchaseOrderItem)
                .join(models.PurchaseOrder)
                .filter(
                    models.PurchaseOrder.sellercloud_po_id == item_data.sellercloud_po_id,
                    models.PurchaseOrderItem.sku == item_data.sku,
                )
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )

        if not item:
            raise HTTPException(status_code=404, detail=f"PO item not found for data: {item_data}")
            
        resolved_items.append((item_data, item))

    sc_items_payload = [
        {
            "PurchaseOrderID": item.purchase_order.sellercloud_po_id if item.purchase_order else item_data.sellercloud_po_id,
            "PurchaseOrderItemID": item.sellercloud_item_id or item_data.sellercloud_item_id,
            "Qty": item_data.qty_in_container,
        }
        for item_data, item in resolved_items
    ]

    return {"Items": sc_items_payload}

# ---------------------------------------------------------------------------
# POST /containers/{container_id}/items
# ---------------------------------------------------------------------------
@router.post("/{container_id}/items")
def add_items_to_container(
    container_id: str,
    items_data: ContainerAddItems,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Add items to an existing container. Syncs to SellerCloud and saves locally.
    """
    container = db.query(models.ShippingContainer).filter(resolve_container_filter(container_id)).first()
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    resolved_items = []
    
    # 1. Resolve and validate items exactly like create_container
    for item_data in items_data.items:
        item = None
        if item_data.po_item_id:
            item = db.query(models.PurchaseOrderItem).filter(models.PurchaseOrderItem.id == item_data.po_item_id).options(joinedload(models.PurchaseOrderItem.purchase_order)).first()
        elif item_data.sellercloud_item_id:
            item = db.query(models.PurchaseOrderItem).filter(models.PurchaseOrderItem.sellercloud_item_id == item_data.sellercloud_item_id).options(joinedload(models.PurchaseOrderItem.purchase_order)).first()
        elif item_data.sellercloud_po_id and item_data.sku:
            item = db.query(models.PurchaseOrderItem).join(models.PurchaseOrder).filter(models.PurchaseOrder.sellercloud_po_id == item_data.sellercloud_po_id, models.PurchaseOrderItem.sku == item_data.sku).options(joinedload(models.PurchaseOrderItem.purchase_order)).first()

        if not item:
            raise HTTPException(status_code=404, detail=f"PO item not found for data: {item_data}")

        # Check availability
        qty_rows = db.query(models.PurchaseOrderItemContainer.qty_in_container).filter(models.PurchaseOrderItemContainer.purchase_order_item_id == item.id).all()
        qty_already_total = sum(r[0] or 0 for r in qty_rows)
        qty_available = max(0, item.qty_ordered - qty_already_total)

        if item_data.qty_in_container > qty_available:
            raise HTTPException(status_code=400, detail=f"SKU '{item.sku}': requested {item_data.qty_in_container} but only {qty_available} available")

        # Check if already in this container
        existing_link = db.query(models.PurchaseOrderItemContainer).filter(models.PurchaseOrderItemContainer.purchase_order_item_id == item.id, models.PurchaseOrderItemContainer.shipping_container_id == container.id).first()
        if existing_link:
            raise HTTPException(status_code=400, detail=f"SKU '{item.sku}' is already in this container. Update functionality is not supported by this endpoint.")

        resolved_items.append((item_data, item))

    # 2. Sync to SellerCloud
    if container.sellercloud_container_id:
        try:
            sc_client = SellerCloudClient()
            sc_items = [
                {
                    "PurchaseOrderID": (item.purchase_order.sellercloud_po_id if item.purchase_order and item.purchase_order.sellercloud_po_id else item_data.sellercloud_po_id),
                    "PurchaseOrderItemID": item.sellercloud_item_id or item_data.sellercloud_item_id,
                    "Qty": item_data.qty_in_container,
                }
                for item_data, item in resolved_items
            ]
            sc_client.add_items_to_container(container.sellercloud_container_id, sc_items)
        except Exception as exc:
            print(f"Failed to add items to SC container {container.sellercloud_container_id}: {exc}")
            raise HTTPException(status_code=500, detail=f"SellerCloud sync failed: {exc}")

    # 3. Update local DB
    linked_items_summary = []
    for item_data, item in resolved_items:
        db.add(
            models.PurchaseOrderItemContainer(
                purchase_order_item_id=item.id,
                shipping_container_id=container.id,
                qty_in_container=item_data.qty_in_container,
                raw_json={"sc_payload_item": item_data.model_dump(mode="json")}
            )
        )
        item.qty_in_container = (item.qty_in_container or 0) + item_data.qty_in_container
        
        linked_items_summary.append({
            "sku": item.sku,
            "qty_added": item_data.qty_in_container,
            "total_item_qty_in_containers": item.qty_in_container
        })

    container.updated_at = datetime.utcnow()
    db.commit()

    # Recalculate PO shipment status for affected POs
    from app.services.po_service import recalculate_po_shipment_status
    affected_po_ids = set(item.purchase_order_id for _, item in resolved_items if item.purchase_order_id)
    for po_id in affected_po_ids:
        recalculate_po_shipment_status(db, str(po_id))

    for item_data, item in resolved_items:
        log_activity(
            db, 
            action="ADD_ITEM_TO_CONTAINER", 
            user_id=current_user.id, 
            entity_type="CONTAINER", 
            entity_id=str(container.id), 
            details={
                "sku": item.sku,
                "qty_added": item_data.qty_in_container,
                "message": f"Added {item_data.qty_in_container} units of {item.sku} to container"
            }
        )

    return {
        "success": True,
        "message": f"Successfully added {len(resolved_items)} items to container",
        "items_added": linked_items_summary
    }


# ---------------------------------------------------------------------------
# POST /containers/{container_id}/activities
# ---------------------------------------------------------------------------
@router.post("/{container_id}/activities")
def add_container_activity(
    container_id: str,
    activity_data: ContainerActivityCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Manually add an activity log/comment to a container.
    """
    container = db.query(models.ShippingContainer).filter(resolve_container_filter(container_id)).first()
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    log_activity(
        db,
        action="ADD_CONTAINER_COMMENT",
        user_id=current_user.id,
        entity_type="CONTAINER",
        entity_id=str(container.id),
        details={"message": activity_data.message}
    )
    return {"success": True, "message": "Activity added successfully"}


# ---------------------------------------------------------------------------
# GET /containers/{container_id}/activities
# ---------------------------------------------------------------------------
@router.get("/{container_id}/activities", response_model=PaginatedResponse)
def get_container_activities(
    container_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Get paginated activity logs for a specific container.
    """
    container = db.query(models.ShippingContainer).filter(resolve_container_filter(container_id)).first()
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    query = db.query(models.UserActivityLog).filter(
        models.UserActivityLog.entity_type == "CONTAINER",
        models.UserActivityLog.entity_id == str(container.id),
        ~models.UserActivityLog.action.ilike("VIEW_%"),
        ~models.UserActivityLog.action.ilike("LIST_%")
    ).order_by(models.UserActivityLog.created_at.desc())

    total = query.count()
    logs = query.offset((page - 1) * page_size).limit(page_size).all()

    from app.services.activity_service import generate_human_readable_message
    results = []
    for log in logs:
        log_out = UserActivityLogOut.model_validate(log)
        if log.user:
            log_out.user_name = log.user.full_name or log.user.email
            
        log_out.human_readable_message = generate_human_readable_message(
            action=log.action,
            entity_type=log.entity_type,
            entity_id=log.entity_id,
            details=log.details,
            user_name=log_out.user_name
        )
            
        results.append(log_out.model_dump(mode='python'))

    return PaginatedResponse(
        total=total,
        page=page,
        page_size=page_size,
        results=results
    )


# ---------------------------------------------------------------------------
# GET /containers/{container_id}/details
# ---------------------------------------------------------------------------
@router.get("/{container_id}/details", response_model=ContainerDetailOut)
def get_container_details(
    container_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Full detail for one container.

    Returns:
    - Container info: name, ETA, received date, `is_received` flag
    - Summary: total qty, item count, unique POs, received qty
    - Items: every SKU in the container with full PO context and received status

    Each item includes `is_fully_received` (True when qty_received >= qty_ordered).
    """
    container = (
        db.query(models.ShippingContainer)
        .filter(resolve_container_filter(container_id))
        .options(
            joinedload(models.ShippingContainer.item_links)
            .joinedload(models.PurchaseOrderItemContainer.item)
            .joinedload(models.PurchaseOrderItem.purchase_order)
            .joinedload(models.PurchaseOrder.vendor),
            joinedload(models.ShippingContainer.attachments),
            joinedload(models.ShippingContainer.warehouse),
            joinedload(models.ShippingContainer.logistics_company),
            joinedload(models.ShippingContainer.comments)
            .joinedload(models.ShippingContainerComment.user),
            joinedload(models.ShippingContainer.comments)
            .joinedload(models.ShippingContainerComment.attachments)
        )
        .first()
    )
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    if current_user.role == "vendor":
        is_authorized = db.query(
            db.query(models.PurchaseOrderItemContainer)
            .join(models.PurchaseOrderItem)
            .join(models.PurchaseOrder)
            .filter(
                models.PurchaseOrderItemContainer.shipping_container_id == container.id,
                models.PurchaseOrder.vendor_id == current_user.vendor_id
            ).exists()
        ).scalar()
        if not is_authorized:
            raise HTTPException(status_code=403, detail="Not authorized to view details for this container")

    items_out = []
    for link in container.item_links:
        item = link.item
        po = item.purchase_order if item else None
        
        # Determine container-specific received quantity
        c_recv = getattr(link, "qty_received_container", None)
        if c_recv is None and link.raw_json and isinstance(link.raw_json, dict):
            c_recv = link.raw_json.get("QtyReceived", 0)
        c_recv = c_recv or 0

        qty_in_cont = link.qty_in_container or 0
        qty_recv_cont = c_recv
        qty_missing_cont = max(0, qty_in_cont - qty_recv_cont)

        items_out.append(
            ContainerDetailItemOut(
                po_item_id=item.id,
                sellercloud_item_id=item.sellercloud_item_id,
                sellercloud_po_id=po.sellercloud_po_id if po else None,
                po_title=po.purchase_title if po else None,
                vendor_name=po.vendor.name if (po and po.vendor) else None,
                sku=item.sku,
                product_name=item.product_name,
                image_url=item.image_url,
                qty_in_container=qty_in_cont,
                qty_received_container=qty_recv_cont,
                qty_missing_container=qty_missing_cont,
                qty_ordered=item.qty_ordered,
                qty_received=item.qty_received,
                qty_remaining=max(0, item.qty_ordered - item.qty_received),
                is_fully_received=item.qty_received >= item.qty_ordered,
                unit_price=float(item.unit_price) if item.unit_price else None,
            )
        )

    # Sort items so that items with container shortage (where qty_received_container < qty_in_container) appear first:
    def item_discrepancy_sort_key(itm: ContainerDetailItemOut):
        qty_in = itm.qty_in_container or 0
        qty_recv = itm.qty_received_container or 0
        shortage = max(0, qty_in - qty_recv)
        has_shortage = shortage > 0
        
        # 1. Shortage items first (0), fully received items last (1)
        # 2. Largest shortage first (-shortage)
        # 3. Alphabetical by SKU
        return (0 if has_shortage else 1, -shortage, itm.sku or "")

    items_out.sort(key=item_discrepancy_sort_key)

    total_qty = sum(i.qty_in_container or 0 for i in items_out)
    total_received_qty = sum(i.qty_received_container or 0 for i in items_out)
    total_missing_qty = max(0, total_qty - total_received_qty)
    unique_po_ids = set(
        str(link.item.purchase_order_id)
        for link in container.item_links
        if link.item
    )
    fully_received_count = sum(1 for i in items_out if i.is_fully_received)

    # Map user names for container comments
    for comment in container.comments:
        if comment.user:
            comment.user_name = comment.user.full_name or comment.user.email

    vendor_credit_comments = [c for c in container.comments if c.category == "vendor_credit"]
    receiving_closure_comments = [c for c in container.comments if c.category == "receiving_closure"]

    status_code, status_label = compute_container_status(
        total_qty_in_container=total_qty,
        total_qty_received=total_received_qty,
        date_emptied=container.date_emptied,
        date_dropped_off=container.date_dropped_off,
        received_date=container.received_date
    )

    return ContainerDetailOut(
        id=container.id,
        sellercloud_container_id=container.sellercloud_container_id,
        container_name=container.container_name,
        estimated_arrival_date=container.estimated_arrival_date,
        received_date=container.received_date,
        is_received=container.received_date is not None,
        container_status=status_code,
        container_status_label=status_label,
        created_at=container.created_at,
        updated_at=container.updated_at,
        warehouse=container.warehouse,
        date_dropped_off=container.date_dropped_off,
        door=container.door,
        trucker_email=container.trucker_email,
        trucking_company=container.trucking_company,
        logistics_company_id=container.logistics_company_id,
        logistics_company=container.logistics_company,
        date_emptied=container.date_emptied,
        unloaded_by=container.unloaded_by,
        unload_cost=float(container.unload_cost) if container.unload_cost is not None else None,
        container_shipping_cost=float(container.container_shipping_cost) if container.container_shipping_cost is not None else None,
        drayage_cost=float(container.drayage_cost) if container.drayage_cost is not None else None,
        customs_duty_misc=float(container.customs_duty_misc) if container.customs_duty_misc is not None else None,
        per_diem=float(container.per_diem) if container.per_diem is not None else None,
        country_of_origin=container.country_of_origin,
        receiving_closure_notes=container.receiving_closure_notes,
        factory_credit_needed=container.factory_credit_needed,
        summary={
            "total_items": len(items_out),
            "container_status": status_code,
            "container_status_label": status_label,
            "total_assigned_quantity": total_qty,
            "total_received_quantity": total_received_qty,
            "missing_quantity": total_missing_qty,
            "total_qty_assigned": total_qty,
            "total_qty_in_container": total_qty,
            "total_qty_received": total_received_qty,
            "total_qty_missing": total_missing_qty,
            "unique_purchase_orders": len(unique_po_ids),
            "fully_received_items": fully_received_count,
            "pending_items": len(items_out) - fully_received_count,
        },
        items=items_out,
        attachments=container.attachments,
        vendor_credit_comments=vendor_credit_comments,
        receiving_closure_comments=receiving_closure_comments
    )


# ---------------------------------------------------------------------------
# POST /containers/{container_id}/attachments
# ---------------------------------------------------------------------------
@router.post("/{container_id}/attachments")
async def add_container_attachments(
    container_id: str,
    files: List[UploadFile] = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Upload one or more files to a container.
    """
    container = db.query(models.ShippingContainer).filter(resolve_container_filter(container_id)).first()
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    allowed_types = [
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/csv"
    ]
    
    # Check sizes and types first
    for f in files:
        if f.size and f.size > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"File {f.filename} exceeds 5MB limit.")
        if f.filename and f.content_type:
            if not f.content_type.startswith('image/') and f.content_type not in allowed_types:
                raise HTTPException(status_code=400, detail=f"File type not allowed for {f.filename}. Only images, PDFs, Word docs, Excel docs, and CSVs are permitted.")
            
    uploaded_attachments = []
    
    for f in files:
        if not f.filename:
            continue
        # Read bytes
        file_bytes = await f.read()
        if len(file_bytes) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"File {f.filename} exceeds 5MB limit.")
            
        file_url = await upload_file_to_gcs(file_bytes, f.filename, f.content_type)
        
        att_model = models.ShippingContainerAttachment(
            shipping_container_id=container.id,
            file_name=f.filename,
            file_url=file_url,
            content_type=f.content_type,
            size=len(file_bytes)
        )
        db.add(att_model)
        uploaded_attachments.append(att_model)
        
    if uploaded_attachments:
        db.commit()
        
    log_activity(db, action="ADD_CONTAINER_ATTACHMENTS", user_id=current_user.id, entity_type="CONTAINER", entity_id=str(container.id), details={"files_uploaded": len(uploaded_attachments)})
    
    return {
        "success": True, 
        "message": f"Successfully uploaded {len(uploaded_attachments)} files",
        "attachments": [
            {
                "id": str(att.id),
                "file_name": att.file_name,
                "file_url": att.file_url,
                "content_type": att.content_type,
                "size": att.size
            } for att in uploaded_attachments
        ]
    }

# ---------------------------------------------------------------------------
# DELETE /containers/attachments/{attachment_id}
# ---------------------------------------------------------------------------
@router.delete("/attachments/{attachment_id}")
def delete_container_attachment(
    attachment_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """
    Delete a container attachment.
    """
    try:
        att_uuid = uuid.UUID(attachment_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid attachment ID format")
        
    attachment = db.query(models.ShippingContainerAttachment).filter(models.ShippingContainerAttachment.id == att_uuid).first()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
        
    container_id = str(attachment.shipping_container_id)
    db.delete(attachment)
    db.commit()
    
    log_activity(db, action="DELETE_CONTAINER_ATTACHMENT", user_id=current_user.id, entity_type="CONTAINER", entity_id=container_id, details={"attachment_id": attachment_id})
    return {"success": True, "message": "Attachment deleted successfully"}



# ---------------------------------------------------------------------------
# POST /containers/{container_id}/sync
# Re-pull container info from SellerCloud
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# POST /containers/{container_id}/sync — Instant Full Container Sync
# ---------------------------------------------------------------------------
@router.post("/{container_id}/sync")
def sync_specific_container(
    container_id: str,
    sync_pos: bool = Query(True, description="Also refresh linked Purchase Orders from SellerCloud"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    **INSTANT SINGLE CONTAINER SYNC**

    Instantly syncs a specific shipping container and its items directly from SellerCloud:
    1. Fetches live container status, dates, and warehouse from SellerCloud.
    2. Discovers all items, `Qty`, and `QtyReceived` inside this container.
    3. If `sync_pos=True`, automatically syncs/refreshes the linked Purchase Orders.
    4. Updates local `shipping_containers` and `purchase_order_item_containers` link tables.
    5. Recalculates PO item quantities and PO shipment statuses.

    **Parameter `{container_id}` accepts:**
    - SellerCloud Container ID (e.g. `5012`)
    - Local Database UUID (e.g. `25af8811-e8e6-47e6-b18a-c7a58741f33c`)
    - Container Name (e.g. `MEDU7337920`)
    """
    from app.services.sync_service import sync_single_container_full
    from app.services.activity_service import log_activity

    try:
        result = sync_single_container_full(db, container_identifier=container_id, sync_pos=sync_pos)
        
        log_activity(
            db,
            action="SYNC_SINGLE_CONTAINER_INSTANT",
            user_id=current_user.id,
            entity_type="SHIPPING_CONTAINER",
            entity_id=str(container_id),
            details={
                "container_name": result.get("container", {}).get("container_name"),
                "sellercloud_container_id": result.get("container", {}).get("sellercloud_container_id"),
                "total_items": result.get("container", {}).get("total_items"),
                "total_qty_in_container": result.get("container", {}).get("total_qty_in_container"),
                "total_qty_received_container": result.get("container", {}).get("total_qty_received_container"),
            }
        )
        return result
    except ValueError as ve:
        raise HTTPException(status_code=404, detail=str(ve))
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error syncing container '{container_id}': {str(e)}")



# ---------------------------------------------------------------------------
# GET /containers/summary/counts
# Quick dashboard counts
# ---------------------------------------------------------------------------
@router.get("/summary/counts")
def get_container_counts(db: Session = Depends(get_db)):
    """
    Quick summary counts for dashboard use.

    Returns:
    - `total` — all containers
    - `received` — containers with received_date set
    - `pending` — containers with no received_date (in transit / not arrived)
    - `synced_to_sc` — containers with a SellerCloud ID
    - `local_only` — containers created locally but SC sync failed (no SC ID)
    """
    total = db.query(models.ShippingContainer).count()
    received = db.query(models.ShippingContainer).filter(
        models.ShippingContainer.received_date.isnot(None)
    ).count()
    synced = db.query(models.ShippingContainer).filter(
        models.ShippingContainer.sellercloud_container_id.isnot(None)
    ).count()

    return {
        "total": total,
        "received": received,
        "pending": total - received,
        "synced_to_sc": synced,
        "local_only": total - synced,
    }


# ---------------------------------------------------------------------------
# GET /containers/by-po/{sellercloud_po_id}
# ---------------------------------------------------------------------------
@router.get("/by-po/{sellercloud_po_id}")
def get_containers_for_po(
    sellercloud_po_id: int,
    db: Session = Depends(get_db),
):
    """
    All containers associated with a specific PO (by SellerCloud PO ID).
    Returns each container grouped with the items from this PO only.
    """
    po = (
        db.query(models.PurchaseOrder)
        .filter(models.PurchaseOrder.sellercloud_po_id == sellercloud_po_id)
        .first()
    )
    if not po:
        raise HTTPException(
            status_code=404, detail=f"Purchase Order {sellercloud_po_id} not found"
        )

    container_links = (
        db.query(models.PurchaseOrderItemContainer)
        .join(
            models.PurchaseOrderItem,
            models.PurchaseOrderItemContainer.purchase_order_item_id
            == models.PurchaseOrderItem.id,
        )
        .filter(models.PurchaseOrderItem.purchase_order_id == po.id)
        .options(
            joinedload(models.PurchaseOrderItemContainer.container),
            joinedload(models.PurchaseOrderItemContainer.item),
        )
        .all()
    )

    containers_dict: dict = {}
    for link in container_links:
        ctr = link.container
        item = link.item
        if ctr.id not in containers_dict:
            containers_dict[ctr.id] = {
                "container": {
                    "id": str(ctr.id),
                    "sellercloud_container_id": ctr.sellercloud_container_id,
                    "container_name": ctr.container_name,
                    "estimated_arrival_date": (
                        ctr.estimated_arrival_date.isoformat()
                        if ctr.estimated_arrival_date
                        else None
                    ),
                    "received_date": (
                        ctr.received_date.isoformat() if ctr.received_date else None
                    ),
                },
                "items": [],
            }
        containers_dict[ctr.id]["items"].append({
            "item_id": str(item.id),
            "sku": item.sku,
            "product_name": item.product_name,
            "qty_in_container": link.qty_in_container,
            "qty_ordered": item.qty_ordered,
            "qty_received": item.qty_received,
            "qty_remaining": item.qty_ordered - item.qty_received,
        })

    containers_list = []
    for cdata in containers_dict.values():
        cdata["total_qty"] = sum(i["qty_in_container"] for i in cdata["items"])
        cdata["items_count"] = len(cdata["items"])
        containers_list.append(cdata)

    return {
        "po_id": str(po.id),
        "sellercloud_po_id": po.sellercloud_po_id,
        "po_title": po.purchase_title,
        "containers_count": len(containers_list),
        "containers": containers_list,
    }


# ---------------------------------------------------------------------------
# POST /containers/debug-sc-create
# Fires the exact SellerCloud payload WITHOUT saving anything locally.
# Use this once to confirm what SC returns so you know the correct ID field.
# ---------------------------------------------------------------------------
@router.post("/debug-sc-create")
def debug_sc_create(
    container_data: ContainerCreate,
    db: Session = Depends(get_db),
):
    """
    **DEBUG ONLY** — sends the container payload to SellerCloud and returns the
    raw response without saving anything to the local database.

    Use this once to answer: *what does SC actually return when a container is
    created?* — specifically which field holds the new container integer ID.

    Once confirmed, the main `POST /containers` endpoint handles it correctly.

    The response includes:
    - `sc_payload_sent` — exactly what we sent to SellerCloud
    - `sc_raw_response` — raw response body as a string
    - `sc_parsed_response` — parsed JSON (or error)
    - `sc_status_code` — HTTP status SC returned
    - `detected_id` — the ID we would extract (if any)
    """
    # Resolve items to build the SC payload (same logic as create, no DB writes)
    sc_items_payload = []
    for item_data in container_data.items:
        item = None
        if item_data.po_item_id:
            item = (
                db.query(models.PurchaseOrderItem)
                .filter(models.PurchaseOrderItem.id == item_data.po_item_id)
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )
        elif item_data.sellercloud_item_id:
            item = (
                db.query(models.PurchaseOrderItem)
                .filter(models.PurchaseOrderItem.sellercloud_item_id == item_data.sellercloud_item_id)
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )
        elif item_data.sellercloud_po_id and item_data.sku:
            item = (
                db.query(models.PurchaseOrderItem)
                .join(models.PurchaseOrder)
                .filter(
                    models.PurchaseOrder.sellercloud_po_id == item_data.sellercloud_po_id,
                    models.PurchaseOrderItem.sku == item_data.sku,
                )
                .options(joinedload(models.PurchaseOrderItem.purchase_order))
                .first()
            )

        if not item:
            raise HTTPException(status_code=404, detail=f"Item not found: {item_data}")

        sc_items_payload.append({
            "PurchaseOrderID": item.purchase_order.sellercloud_po_id if item.purchase_order else item_data.sellercloud_po_id,
            "PurchaseOrderItemID": item.sellercloud_item_id or item_data.sellercloud_item_id,
            "Qty": item_data.qty_in_container,
        })

    # Per SC Swagger: POST /api/ShippingContainers only accepts name/dates
    # Items go to POST /api/ShippingContainers/{id}/Items separately
    sc_payload = {
        "ContainerName": container_data.container_name,
        "EstimatedArrivalDate": container_data.estimated_arrival_date.isoformat() if container_data.estimated_arrival_date else None,
        "ShippingStatus": 2 if container_data.received_date else 1,
    }
    if container_data.received_date:
        sc_payload["ReceivedDate"] = container_data.received_date.isoformat()

    # Build the add-items payload for display (not called since this is debug-only)
    sc_add_items_payload = {"Items": sc_items_payload}

    sc_raw = None
    sc_parsed = None
    sc_status = None
    sc_items_raw = None
    sc_items_status = None
    error = None

    try:
        sc_client = SellerCloudClient()
        # STEP 1: Test the container create call
        resp = sc_client._request("POST", "/api/ShippingContainers", json=sc_payload)
        sc_status = resp.status_code
        sc_raw = resp.text
        try:
            sc_parsed = resp.json()
        except Exception:
            sc_parsed = {"_not_json": sc_raw}

        # Extract container ID from the create response
        detected_id = None
        if isinstance(sc_parsed, dict):
            detected_id = (
                sc_parsed.get("ID") or sc_parsed.get("Id") or
                sc_parsed.get("ContainerID") or sc_parsed.get("ContainerId") or
                sc_parsed.get("id")
            )
        elif isinstance(sc_parsed, int):
            detected_id = sc_parsed
        # Also handle plain integer body
        if not detected_id and sc_raw and sc_raw.strip().isdigit():
            detected_id = int(sc_raw.strip())

        # STEP 2: If we got an ID, test the add-items call too
        if detected_id and sc_items_payload:
            try:
                items_resp = sc_client._request(
                    "POST",
                    f"/api/ShippingContainers/{detected_id}/Items",
                    json={"Items": sc_items_payload},
                )
                sc_items_status = items_resp.status_code
                sc_items_raw = items_resp.text
            except Exception as items_exc:
                sc_items_raw = f"ERROR: {items_exc}"

    except Exception as exc:
        error = str(exc)
        detected_id = None

    return {
        "note": "DEBUG ONLY — this DID create a container + add items in SellerCloud. Delete it from SC if testing.",
        "step1_create_container": {
            "sc_status_code": sc_status,
            "sc_payload_sent": sc_payload,
            "sc_raw_response": sc_raw,
            "sc_parsed_response": sc_parsed,
            "detected_id": detected_id,
        },
        "step2_add_items": {
            "sc_status_code": sc_items_status,
            "sc_payload_sent": sc_add_items_payload if detected_id else "SKIPPED — no container ID from step 1",
            "sc_raw_response": sc_items_raw,
        },
        "error": error,
    }

import csv
import io
from fastapi.responses import StreamingResponse
from app.schemas import ContainerExportRequest
from sqlalchemy.orm import joinedload

# ---------------------------------------------------------------------------
# POST /containers/export/csv
# Export containers to CSV
# ---------------------------------------------------------------------------
@router.post("/export/csv")
def export_containers_csv(
    request_data: ContainerExportRequest,
    db: Session = Depends(get_db)
):
    """
    Export containers to a CSV file.
    Creates a row for every item inside the matched containers.
    """
    query = db.query(models.ShippingContainer)
    
    if request_data.container_ids:
        query = query.filter(models.ShippingContainer.id.in_(request_data.container_ids))
        
    if request_data.is_received is not None:
        if request_data.is_received:
            query = query.filter(models.ShippingContainer.received_date.isnot(None))
        else:
            query = query.filter(models.ShippingContainer.received_date.is_(None))
            
    # Eager load relationships
    query = query.options(
        joinedload(models.ShippingContainer.warehouse),
        joinedload(models.ShippingContainer.item_links)
        .joinedload(models.PurchaseOrderItemContainer.item)
        .joinedload(models.PurchaseOrderItem.purchase_order)
    )
    
    containers = query.order_by(models.ShippingContainer.created_at.desc()).all()
    
    # Define available columns
    DEFAULT_COLUMNS = [
        "container_name",
        "sellercloud_container_id",
        "estimated_arrival_date",
        "received_date",
        "warehouse_name",
        "sellercloud_po_id",
        "po_order_number",
        "sku",
        "item_name",
        "qty_ordered",
        "qty_in_container"
    ]
    
    columns = request_data.columns if request_data.columns else DEFAULT_COLUMNS
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columns)
    
    item_specific_cols = {
        "sellercloud_po_id",
        "po_order_number",
        "sku",
        "item_name",
        "qty_ordered",
        "qty_in_container"
    }
    
    requires_items = any(col in item_specific_cols for col in columns)
    
    for ctr in containers:
        warehouse_name = ctr.warehouse.name if ctr.warehouse else ""
        
        if not ctr.item_links or not requires_items:
            # Container with no items or item-level columns not requested
            row_dict = {
                "container_name": ctr.container_name or "",
                "sellercloud_container_id": ctr.sellercloud_container_id or "",
                "estimated_arrival_date": ctr.estimated_arrival_date.strftime("%Y-%m-%d") if ctr.estimated_arrival_date else "",
                "received_date": ctr.received_date.strftime("%Y-%m-%d") if ctr.received_date else "",
                "warehouse_name": warehouse_name,
                "sellercloud_po_id": "",
                "po_order_number": "",
                "sku": "",
                "item_name": "",
                "qty_ordered": "",
                "qty_in_container": ""
            }
            writer.writerow([row_dict.get(col, "") for col in columns])
        else:
            for link in ctr.item_links:
                item = link.item
                po = item.purchase_order if item else None
                
                po_order_number = ""
                if po and po.purchase_title:
                    if "cloned from po" in po.purchase_title.lower():
                        po_order_number = "Stock"
                    else:
                        import re
                        match = re.search(r'#(\d+)', po.purchase_title)
                        if match:
                            po_order_number = match.group(1)

                row_dict = {
                    "container_name": ctr.container_name or "",
                    "sellercloud_container_id": ctr.sellercloud_container_id or "",
                    "estimated_arrival_date": ctr.estimated_arrival_date.strftime("%Y-%m-%d") if ctr.estimated_arrival_date else "",
                    "received_date": ctr.received_date.strftime("%Y-%m-%d") if ctr.received_date else "",
                    "warehouse_name": warehouse_name,
                    "sellercloud_po_id": po.sellercloud_po_id if po else "",
                    "po_order_number": po_order_number,
                    "sku": item.sku if item else "",
                    "item_name": item.product_name if item else "",
                    "qty_ordered": item.qty_ordered if item else "",
                    "qty_in_container": link.qty_in_container or 0
                }
                writer.writerow([row_dict.get(col, "") for col in columns])
                
    output.seek(0)
    from datetime import datetime
    filename = f"containers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ---------------------------------------------------------------------------
# POST /containers/import/preview
# Read an uploaded CSV or Excel file and return found PO items
# ---------------------------------------------------------------------------
@router.post("/import/preview")
async def preview_container_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Reads an uploaded CSV or Excel file containing container items.
    Tries to map each row to a PurchaseOrderItem in the database.
    
    Expected columns (flexible naming):
    - POID / PO ID
    - ProductID / SKU / ProductID - SKU
    - Quantity / Qty
    
    Returns a JSON array where each row shows the data from the file
    plus the matched `po_item_id` and item details, or `null` if not found.
    """
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only CSV or Excel files are supported")
        
    try:
        import io
        import csv
        
        contents = await file.read()
        rows = []
        headers = []
        
        if file.filename.endswith('.csv'):
            text_content = contents.decode('utf-8-sig', errors='replace')
            csv_reader = csv.reader(io.StringIO(text_content))
            try:
                headers = [str(h).strip() for h in next(csv_reader)]
                for row_data in csv_reader:
                    # pad row to match headers length
                    row_data = row_data + [''] * (len(headers) - len(row_data))
                    rows.append(dict(zip(headers, row_data)))
            except StopIteration:
                pass
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = wb.active
            
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row:
                headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]
                for row_data in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = list(row_data) + [''] * (len(headers) - len(row_data))
                    row_data = [str(cell) if cell is not None else "" for cell in row_data]
                    rows.append(dict(zip(headers, row_data)))
                    
        # Identify columns
        po_col = next((c for c in headers if 'po' in c.lower() and 'id' in c.lower()), None)
        if not po_col:
            po_col = next((c for c in headers if c.lower() in ['po', 'po number', 'purchase order']), None)
            
        sku_col = next((c for c in headers if 'sku' in c.lower() or 'productid' in c.lower() or 'product id' in c.lower()), None)
        qty_col = next((c for c in headers if 'qty' in c.lower() or 'quantity' in c.lower()), None)
        
        if not po_col or not sku_col:
            raise HTTPException(status_code=400, detail=f"Could not identify PO ID or SKU columns. Found columns: {', '.join(headers)}")
            
        results = []
        
        for index, row in enumerate(rows):
            po_val = str(row.get(po_col, "")).strip()
            sku_val = str(row.get(sku_col, "")).strip()
            
            qty_val = 0
            raw_qty = row.get(qty_col)
            if raw_qty:
                try:
                    qty_val = int(float(raw_qty))
                except ValueError:
                    qty_val = 0
                    
            if not po_val or not sku_val:
                continue
                
            # Clean up PO value in case it has # or .0
            import re
            po_val_clean = re.sub(r'[^0-9]', '', po_val)
            if not po_val_clean:
                po_val_clean = po_val
                
            # Lookup PO item in DB
            try:
                sellercloud_po_id = int(po_val_clean)
            except ValueError:
                sellercloud_po_id = -1
                
            # Check PO existence first
            po = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.sellercloud_po_id == sellercloud_po_id).first()
            po_item = None
            validation_message = "Valid"
            status = "success"
            
            if not po:
                validation_message = f"PO {sellercloud_po_id} not found in database. Try syncing it first."
                status = "error"
            else:
                # PO exists, check SKU
                po_item = (
                    db.query(models.PurchaseOrderItem)
                    .filter(
                        models.PurchaseOrderItem.purchase_order_id == po.id,
                        models.PurchaseOrderItem.sku.ilike(f"%{sku_val}%")
                    )
                    .first()
                )
                if not po_item:
                    validation_message = f"SKU '{sku_val}' not found in PO {sellercloud_po_id}."
                    status = "error"
            
            row_result = {
                "row_index": index + 1,
                "file_po_id": po_val,
                "file_sku": sku_val,
                "file_qty": qty_val,
                "status": status,
                "validation_message": validation_message,
                "found_item": None
            }
            
            if po_item:
                qty_already = sum(link.qty_in_container or 0 for link in po_item.container_links) if po_item.container_links else 0
                qty_available = max(0, po_item.qty_ordered - qty_already)
                
                row_result["found_item"] = {
                    "po_item_id": str(po_item.id),
                    "purchase_order_id": str(po_item.purchase_order_id),
                    "sellercloud_item_id": po_item.sellercloud_item_id,
                    "product_name": po_item.product_name,
                    "qty_ordered": po_item.qty_ordered,
                    "qty_received": po_item.qty_received,
                    "qty_already_in_containers": qty_already,
                    "qty_available_for_container": qty_available
                }
                
            results.append(row_result)
            
        return {
            "success": True,
            "message": f"Processed {len(results)} rows",
            "columns_identified": {
                "po_column": po_col,
                "sku_column": sku_col,
                "quantity_column": qty_col
            },
            "data": results
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@router.get("/validate-item")
def validate_container_item(
    po_id: str = Query(..., description="The SellerCloud PO ID"),
    sku: str = Query(..., description="The SKU of the item"),
    db: Session = Depends(get_db)
):
    """
    Validates a single PO/SKU pair for a manual row entry on the frontend.
    Returns exactly the same validation structure as the CSV import preview.
    """
    import re
    
    # Clean up PO value in case it has # or .0
    po_val_clean = re.sub(r'[^0-9]', '', str(po_id))
    if not po_val_clean:
        po_val_clean = str(po_id)
        
    try:
        sellercloud_po_id = int(po_val_clean)
    except ValueError:
        sellercloud_po_id = -1
        
    po = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.sellercloud_po_id == sellercloud_po_id).first()
    po_item = None
    validation_message = "Valid"
    status = "success"
    
    if not po:
        validation_message = f"PO {sellercloud_po_id} not found in database. Try syncing it first."
        status = "error"
    else:
        # PO exists, check SKU
        po_item = (
            db.query(models.PurchaseOrderItem)
            .filter(
                models.PurchaseOrderItem.purchase_order_id == po.id,
                models.PurchaseOrderItem.sku.ilike(f"%{sku}%")
            )
            .first()
        )
        if not po_item:
            validation_message = f"SKU '{sku}' not found in PO {sellercloud_po_id}."
            status = "error"
            
    result = {
        "file_po_id": str(po_id),
        "file_sku": str(sku),
        "status": status,
        "validation_message": validation_message,
        "found_item": None
    }
    
    if po_item:
        qty_already = sum(link.qty_in_container or 0 for link in po_item.container_links) if po_item.container_links else 0
        qty_available = max(0, po_item.qty_ordered - qty_already)
        
        result["found_item"] = {
            "po_item_id": str(po_item.id),
            "purchase_order_id": str(po_item.purchase_order_id),
            "sellercloud_item_id": po_item.sellercloud_item_id,
            "product_name": po_item.product_name,
            "qty_ordered": po_item.qty_ordered,
            "qty_received": po_item.qty_received,
            "qty_already_in_containers": qty_already,
            "qty_available_for_container": qty_available
        }
        
    return result

@router.post("/validate-items-bulk")
def validate_container_items_bulk(
    request: schemas.ValidateContainerBulkRequest,
    db: Session = Depends(get_db)
):
    """
    Validates multiple PO/SKU pairs for manual row entries on the frontend.
    Returns exactly the same validation structure as the CSV import preview for each row.
    """
    import re
    results = []
    
    for index, item in enumerate(request.items):
        po_id = item.po_id
        sku = item.sku
        qty = item.qty
        
        # Clean up PO value in case it has # or .0
        po_val_clean = re.sub(r'[^0-9]', '', str(po_id))
        if not po_val_clean:
            po_val_clean = str(po_id)
            
        try:
            sellercloud_po_id = int(po_val_clean)
        except ValueError:
            sellercloud_po_id = -1
            
        po = db.query(models.PurchaseOrder).filter(models.PurchaseOrder.sellercloud_po_id == sellercloud_po_id).first()
        po_item = None
        validation_message = "Valid"
        status = "success"
        
        if not po:
            validation_message = f"PO {sellercloud_po_id} not found in database. Try syncing it first."
            status = "error"
        else:
            # PO exists, check SKU
            po_item = (
                db.query(models.PurchaseOrderItem)
                .filter(
                    models.PurchaseOrderItem.purchase_order_id == po.id,
                    models.PurchaseOrderItem.sku.ilike(f"%{sku}%")
                )
                .first()
            )
            if not po_item:
                validation_message = f"SKU '{sku}' not found in PO {sellercloud_po_id}."
                status = "error"
                
        row_result = {
            "row_index": index + 1,
            "file_po_id": str(po_id),
            "file_sku": str(sku),
            "file_qty": qty,
            "status": status,
            "validation_message": validation_message,
            "found_item": None
        }
        
        if po_item:
            qty_already = sum(link.qty_in_container or 0 for link in po_item.container_links) if po_item.container_links else 0
            qty_available = max(0, po_item.qty_ordered - qty_already)
            
            row_result["found_item"] = {
                "po_item_id": str(po_item.id),
                "purchase_order_id": str(po_item.purchase_order_id),
                "sellercloud_item_id": po_item.sellercloud_item_id,
                "product_name": po_item.product_name,
                "qty_ordered": po_item.qty_ordered,
                "qty_received": po_item.qty_received,
                "remaining_qty": max(0, po_item.qty_ordered - (po_item.qty_received or 0)),
                "qty_already_in_containers": qty_already,
                "qty_available_for_container": qty_available
            }
            
        results.append(row_result)
        
    return {
        "success": True,
        "message": f"Validated {len(results)} rows",
        "data": results
    }

@router.get("/allways/search/{container_number}")
def search_container_allways(
    container_number: str,
    db: Session = Depends(get_db)
):
    """
    Search AllWays API for tracking information for a specific container.
    """
    from app.services.allways_service import track_container
    try:
        result = track_container(container_number)
        if result.get("error_message") and not result.get("raw_response"):
            raise HTTPException(status_code=404, detail=result["error_message"])
            
        # Add warehouse info based on AllWays destination port
        result["warehouse_id"] = None
        result["warehouse_name"] = None
        
        dest_str = result.get("destination_port", "")
        warehouse_name_query = None
        
        if dest_str:
            dest_str_upper = dest_str.upper()
            if "NEW YORK" in dest_str_upper or " NY" in dest_str_upper or dest_str_upper == "NY":
                warehouse_name_query = "South Brunswick"
            elif "LOS ANGELES" in dest_str_upper or " CA" in dest_str_upper or dest_str_upper == "CA":
                warehouse_name_query = "California"
                
        if warehouse_name_query:
            warehouse = db.query(models.Warehouse).filter(
                models.Warehouse.name.ilike(f"%{warehouse_name_query}%")
            ).first()
            if warehouse:
                result["warehouse_id"] = str(warehouse.id)
                result["warehouse_name"] = warehouse.name
                
        # Fallback to existing container in DB if AllWays didn't have a mapped port
        if not result["warehouse_id"]:
            container = db.query(models.ShippingContainer).filter(models.ShippingContainer.container_name == container_number).first()
            if container and container.warehouse_id:
                result["warehouse_id"] = str(container.warehouse_id)
                if container.warehouse:
                    result["warehouse_name"] = container.warehouse.name
                
        # Remove raw_response to avoid cluttering the payload
        result.pop("raw_response", None)
        return {"success": True, "data": result}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/{container_id}/tracking/sync")
def sync_single_container_tracking(
    container_id: str,
    db: Session = Depends(get_db)
):
    """
    Sync tracking information for a single container with AllWays API.
    Accepts either the internal UUID or the container number (e.g., TLLU5203239).
    """
    import uuid
    container = None
    try:
        # Try to parse as UUID
        val = uuid.UUID(container_id)
        container = db.query(models.ShippingContainer).filter(models.ShippingContainer.id == val).first()
    except ValueError:
        # Fallback to searching by container_name (container number)
        container = db.query(models.ShippingContainer).filter(models.ShippingContainer.container_name == container_id).first()

    if not container:
        raise HTTPException(status_code=404, detail=f"Container '{container_id}' not found in database")
        
    if not container.container_name:
        raise HTTPException(status_code=400, detail="Container has no container number assigned")
        
    from app.services.allways_service import sync_container_tracking
    try:
        tracking = sync_container_tracking(db, container)
        if tracking.error_message and not tracking.origin_port and not tracking.latitude:
            return {"success": False, "message": f"Sync failed: {tracking.error_message}"}
        
        # Prepare tracking info for response
        tracking_info = {
            "status": tracking.status, 
            "eta": tracking.eta.isoformat() if tracking.eta else None,
            "etd": tracking.etd.isoformat() if tracking.etd else None,
            "carrier": tracking.carrier,
            "origin_port": tracking.origin_port,
            "destination_port": tracking.destination_port
        }
        return {"success": True, "message": "Container tracking synced successfully", "data": tracking_info}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/tracking/sync-all")
def trigger_allways_sync_all(background_tasks: BackgroundTasks):
    """
    Triggers a background job to sync all active containers with AllWays API.
    """
    def background_sync_all_tracking():
        print("Starting AllWays background sync...", flush=True)
        from app.database import SessionLocal
        db = SessionLocal()
        try:
            from app.services.allways_service import sync_all_containers_tracking
            sync_all_containers_tracking(db)
            print("Finished AllWays background sync.", flush=True)
        except Exception as e:
            print(f"CRITICAL ERROR in AllWays background sync: {e}", flush=True)
        finally:
            db.close()
            
    background_tasks.add_task(background_sync_all_tracking)
    return {"success": True, "message": "AllWays container sync started in the background."}


# ---------------------------------------------------------------------------
# CONTAINER COMMENTS ENDPOINTS
# ---------------------------------------------------------------------------

@router.post("/{container_id}/comments", response_model=ShippingContainerCommentOut)
async def add_container_comment(
    container_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    comment: Optional[str] = Form(None),
    category: str = Form(...), # 'vendor_credit' or 'receiving_closure'
    parent_id: Optional[str] = Form(None),
    tagged_user_ids: Optional[str] = Form(None),
    files: list[UploadFile] = File([]),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    import uuid
    import json
    from app.config import settings
    from app.services.gcs_service import upload_file_to_gcs

    container = db.query(models.ShippingContainer).filter(resolve_container_filter(container_id)).first()
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    if current_user.role == "vendor":
        is_authorized = db.query(
            db.query(models.PurchaseOrderItemContainer)
            .join(models.PurchaseOrderItem)
            .join(models.PurchaseOrder)
            .filter(
                models.PurchaseOrderItemContainer.shipping_container_id == container.id,
                models.PurchaseOrder.vendor_id == current_user.vendor_id
            ).exists()
        ).scalar()
        if not is_authorized:
            raise HTTPException(status_code=403, detail="Not authorized to view details for this container")

    if not comment:
        try:
            body = await request.json()
            comment = body.get("comment")
            category = body.get("category", category)
            parent_id = body.get("parent_id")
            tagged_users_list = body.get("tagged_user_ids", [])
            if isinstance(tagged_users_list, list):
                tagged_user_ids = json.dumps(tagged_users_list)
            else:
                tagged_user_ids = str(tagged_users_list)
        except Exception:
            pass

    if not comment:
        if files:
            comment = ""
        else:
            raise HTTPException(status_code=400, detail="comment field is required")

    try:
        if not tagged_user_ids or tagged_user_ids == "null":
            tagged_users = []
        else:
            tagged_users = json.loads(tagged_user_ids)
            if not isinstance(tagged_users, list):
                tagged_users = [tagged_users]
    except Exception:
        tagged_users = [uid.strip() for uid in tagged_user_ids.split(",") if uid.strip()]

    if parent_id == "null" or not parent_id:
        parent_uuid = None
    else:
        try:
            parent_uuid = uuid.UUID(parent_id)
        except ValueError:
            parent_uuid = None

    new_comment = models.ShippingContainerComment(
        shipping_container_id=container.id,
        user_id=current_user.id,
        comment=comment,
        category=category,
        parent_id=parent_uuid
    )
    db.add(new_comment)
    db.flush()
    db.refresh(new_comment)
    new_comment.user_name = current_user.full_name or current_user.email
    
    uploaded_attachments = []
    email_attachments = []
    
    allowed_types = [
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/csv"
    ]
    
    for f in files:
        if f.size and f.size > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"File {f.filename} exceeds 5MB limit.")
        if f.filename and f.content_type:
            if not f.content_type.startswith('image/') and f.content_type not in allowed_types:
                raise HTTPException(status_code=400, detail=f"File type not allowed for {f.filename}. Only images, PDFs, Word docs, Excel docs, and CSVs are permitted.")
            
    for f in files:
        if not f.filename:
            continue
        file_bytes = await f.read()
        if len(file_bytes) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"File {f.filename} exceeds 5MB limit.")
            
        file_url = await upload_file_to_gcs(file_bytes, f.filename, f.content_type)
        
        att_model = models.ShippingContainerCommentAttachment(
            comment_id=new_comment.id,
            file_name=f.filename,
            file_url=file_url,
            content_type=f.content_type,
            size=len(file_bytes)
        )
        db.add(att_model)
        uploaded_attachments.append(att_model)
        email_attachments.append({
            "file_name": f.filename,
            "content": file_bytes,
            "content_type": f.content_type
        })
        
    db.commit()
    db.refresh(new_comment)
    
    # Process Tags and Mentions
    section_name = "Vendor Credit Needed" if category == "vendor_credit" else "Receiving Closure Notes"
    link = f"{settings.FRONTEND_ORIGIN}/containers/{container.id}?comment_id={new_comment.id}"
    
    await process_container_comment_tags(
        db=db,
        tagged_user_ids=tagged_users,
        commenter_name=new_comment.user_name,
        link=link,
        background_tasks=background_tasks,
        is_edit=False,
        section=section_name,
        container_name=container.container_name,
        comment_text=new_comment.comment,
        attachments=email_attachments
    )
    
    log_activity(db, action="ADD_CONTAINER_COMMENT", user_id=current_user.id, entity_type="CONTAINER", entity_id=str(container.id), details={"comment_id": str(new_comment.id), "category": category})
    return new_comment


@router.put("/comments/{comment_id}", response_model=ShippingContainerCommentOut)
async def update_container_comment(
    comment_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    comment_text: Optional[str] = Form(None, alias="comment"),
    tagged_user_ids: Optional[str] = Form(None),
    files: list[UploadFile] = File([]),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    import uuid
    import json
    from app.config import settings
    from app.services.gcs_service import upload_file_to_gcs

    comment = db.query(models.ShippingContainerComment).filter(models.ShippingContainerComment.id == comment_id).first()
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")
    if comment.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this comment")
        
    container = comment.container
    if comment_text is not None:
        comment.comment = comment_text
        comment.is_edited = True
        
    tagged_users = []
    if tagged_user_ids:
        try:
            tagged_users = json.loads(tagged_user_ids)
            if not isinstance(tagged_users, list):
                tagged_users = [tagged_users]
        except Exception:
            tagged_users = [uid.strip() for uid in tagged_user_ids.split(",") if uid.strip()]
            
    uploaded_attachments = []
    email_attachments = []
    
    allowed_types = [
        "application/pdf",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/csv"
    ]
    
    for f in files:
        if f.size and f.size > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"File {f.filename} exceeds 5MB limit.")
        if f.filename and f.content_type:
            if not f.content_type.startswith('image/') and f.content_type not in allowed_types:
                raise HTTPException(status_code=400, detail=f"File type not allowed for {f.filename}. Only images, PDFs, Word docs, Excel docs, and CSVs are permitted.")
                
    for f in files:
        if not f.filename:
            continue
        file_bytes = await f.read()
        if len(file_bytes) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"File {f.filename} exceeds 5MB limit.")
            
        file_url = await upload_file_to_gcs(file_bytes, f.filename, f.content_type)
        
        att_model = models.ShippingContainerCommentAttachment(
            comment_id=comment.id,
            file_name=f.filename,
            file_url=file_url,
            content_type=f.content_type,
            size=len(file_bytes)
        )
        db.add(att_model)
        uploaded_attachments.append(att_model)
        email_attachments.append({
            "file_name": f.filename,
            "content": file_bytes,
            "content_type": f.content_type
        })
        
    db.commit()
    db.refresh(comment)
    comment.user_name = current_user.full_name or current_user.email
    
    # Process Tags
    section_name = "Vendor Credit Needed" if comment.category == "vendor_credit" else "Receiving Closure Notes"
    link = f"{settings.FRONTEND_ORIGIN}/containers/{container.id}?comment_id={comment.id}"
    
    await process_container_comment_tags(
        db=db,
        tagged_user_ids=tagged_users,
        commenter_name=comment.user_name,
        link=link,
        background_tasks=background_tasks,
        is_edit=True,
        section=section_name,
        container_name=container.container_name if container else None,
        comment_text=comment.comment,
        attachments=email_attachments
    )
    
    log_activity(db, action="UPDATE_CONTAINER_COMMENT", user_id=current_user.id, entity_type="CONTAINER", entity_id=str(container.id) if container else None, details={"comment_id": comment_id})
    return comment


@router.delete("/comments/{comment_id}")
def delete_container_comment(
    comment_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    comment = db.query(models.ShippingContainerComment).filter(models.ShippingContainerComment.id == comment_id).first()
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")
    if comment.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this comment")
        
    container = comment.container
    from app.services.gcs_service import delete_file_from_gcs
    for attachment in comment.attachments:
        if attachment.file_url:
            background_tasks.add_task(delete_file_from_gcs, attachment.file_url)
            
    db.delete(comment)
    db.commit()
    
    log_activity(db, action="DELETE_CONTAINER_COMMENT", user_id=current_user.id, entity_type="CONTAINER", entity_id=str(container.id) if container else None, details={"comment_id": comment_id})
    return {"success": True, "message": "Comment deleted successfully"}


@router.delete("/comments/attachments/{attachment_id}")
def delete_container_comment_attachment(
    attachment_id: str,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    attachment = db.query(models.ShippingContainerCommentAttachment).filter(models.ShippingContainerCommentAttachment.id == attachment_id).first()
    if not attachment:
        raise HTTPException(status_code=404, detail="Attachment not found")
        
    comment = attachment.comment
    if comment and comment.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this attachment")
        
    from app.services.gcs_service import delete_file_from_gcs
    if attachment.file_url:
        background_tasks.add_task(delete_file_from_gcs, attachment.file_url)
        
    db.delete(attachment)
    db.commit()
    
    return {"success": True, "message": "Attachment deleted successfully"}


@router.get("/{container_id}/comments", response_model=List[ShippingContainerCommentOut])
def get_container_comments(
    container_id: str,
    category: Optional[str] = Query(None, description="Filter comments by category: 'vendor_credit' or 'receiving_closure'"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    container = db.query(models.ShippingContainer).filter(resolve_container_filter(container_id)).first()
    if not container:
        raise HTTPException(status_code=404, detail="Container not found")

    if current_user.role == "vendor":
        is_authorized = db.query(
            db.query(models.PurchaseOrderItemContainer)
            .join(models.PurchaseOrderItem)
            .join(models.PurchaseOrder)
            .filter(
                models.PurchaseOrderItemContainer.shipping_container_id == container.id,
                models.PurchaseOrder.vendor_id == current_user.vendor_id
            ).exists()
        ).scalar()
        if not is_authorized:
            raise HTTPException(status_code=403, detail="Not authorized to view comments for this container")

    query = db.query(models.ShippingContainerComment).filter(
        models.ShippingContainerComment.shipping_container_id == container.id
    )

    if category:
        query = query.filter(models.ShippingContainerComment.category == category)

    comments = query.options(
        joinedload(models.ShippingContainerComment.user),
        joinedload(models.ShippingContainerComment.attachments)
    ).order_by(models.ShippingContainerComment.created_at.asc()).all()

    for comment in comments:
        if comment.user:
            comment.user_name = comment.user.full_name or comment.user.email

    return comments


async def process_container_comment_tags(
    db, tagged_user_ids, commenter_name, link, background_tasks, 
    is_edit=False, section="Containers", container_name=None, 
    comment_text="", attachments=None
):
    import app.models as models
    import re
    from app.services.email_service import send_tag_notification
    
    explicit_ids = set()
    if tagged_user_ids:
        found_uuids = re.findall(r'[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}', str(tagged_user_ids))
        explicit_ids.update(found_uuids)

    if comment_text:
        matches = re.findall(r'@([A-Za-z0-9_\-\.]+)', comment_text)
        if matches:
            users = db.query(models.User).all()
            for match in matches:
                match_clean = match.lower().replace("_", " ")
                for user in users:
                    full_name = (user.full_name or "").lower()
                    first_name = (user.first_name or "").lower()
                    if full_name and match_clean in full_name:
                        explicit_ids.add(str(user.id))
                    elif first_name and match_clean in first_name:
                        explicit_ids.add(str(user.id))

    if not explicit_ids:
        return
        
    users = db.query(models.User).filter(models.User.id.in_(list(explicit_ids))).all()
    emails = [u.email for u in users if u.email]
    if emails:
        background_tasks.add_task(
            send_tag_notification, 
            emails=emails, 
            commenter_name=commenter_name, 
            link=link, 
            is_edit=is_edit, 
            section=section,
            po_number=None,
            sku=None,
            comment_text=comment_text,
            attachments=attachments,
            container_name=container_name
        )




